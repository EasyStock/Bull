import pandas as pd



from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.utils import column_index_from_string
from Utility.convertDataFrameToJPG import DataFrameToJPG
from workspace import workSpaceRoot,GetFuPanRoot,GetStockFolder
import os

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )

class CVolumnSelectorToXLSFormatter(object):
    def __init__(self):
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = "爆量复盘"
        self.title = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 14
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12
        sheet.column_dimensions['I'].width = 12
        sheet.column_dimensions['J'].width = 16
        sheet.column_dimensions['K'].width = 14
        sheet.column_dimensions['L'].width = 14

    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","L",70)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 2

    def AddHeader(self,sheet,header):
        self.mergeRow(sheet,self.rows+1,"A","L",30)
        cell = sheet.cell(self.rows+1,1)
        cell.alignment = alignment_left
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  header
        cell.font = Font(name='宋体', size=18, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddData(self,excelWriter,df:pd.DataFrame):
        df.to_excel(excelWriter, sheet_name= self.sheetName,index=False,startrow=self.rows)
        sheet = excelWriter.sheets[self.sheetName]
        startRow = self.rows  + 1
        endRow = startRow + df.shape[0] + 1
        for index in range(startRow,endRow):
            for column in range(1,13):
                cell = sheet.cell(row = index, column = column)
                cell.border = border
                cell.alignment = alignment_center
                cell.font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)
                if index % 2 != 0:
                    cell.fill = PatternFill('solid', fgColor="CCEEFF")
                else:
                    cell.fill = PatternFill('solid', fgColor="FFFFFF")
                
        self.rows = self.rows + df.shape[0] + 3

class CVMASelecter(object):
    def __init__(self,dbConnection):
        self.dbConnection = dbConnection

    def _Select(self,stockID,stockName,date,zhangDiefu,VMA,VMAValue):
        sql = f'''SELECT * from  `stock`.`stockdailyinfo_traning_result` where `stockID` = "{stockID}" and VMA = "{VMA}";'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        df["涨幅"] = df['涨幅'].astype(float)
        df["VMA值"] = df['VMA值'].astype(float)
        df["平均涨幅"] = df['平均涨幅'].astype(float)
        df["仓位"] = df['仓位'].astype(float)
        df = df[df['VMA值']<=VMAValue]
        df = df[df['涨幅']<=zhangDiefu]
        df["股票简称"] = stockName
        df["日期"] = date
        df["今日VMA值"] = VMAValue
        df["今日涨跌幅"] = zhangDiefu
        if df.empty:
            return (False,None)
        
        flag1 = (df['几日后涨幅'] == "1日后涨幅") & (df['平均涨幅']>=1)
        flag2 = (df['几日后涨幅'] == "3日后涨幅") & (df['平均涨幅']>=1)
        flag3 = (df['几日后涨幅'] == "5日后涨幅") & (df['平均涨幅']>=1)
        flag4 = (df['几日后涨幅'] == "7日后涨幅") & (df['平均涨幅']>=1)
        df = df[flag1 | flag2 | flag3 | flag4]
        df = df[df["仓位"] > 0]
        if df.empty:
            return (False,None)
        
        df.reset_index(drop=True,inplace=True)
        message = f'''股票代码:{stockID} 股票简称:{stockName} 日期:{date} VMA值:<= {VMAValue} 涨跌幅<={zhangDiefu}有:'''
        return (True,(df,stockID,stockName,message))

    
    def Select(self,date,VMA = 60, vmaThreshold = 2):
        #sql = f'''SELECT A.*,B.`股票简称` FROM `stockdailyinfo` As A,`stockbasicinfo` As B where A.`日期` = "{date}" and A.`股票代码` = B.`股票代码`and A.`V/MA60` >= 2;'''
        key = f'''V/MA{VMA}'''
        sql = f'''SELECT `日期`,`股票代码`,`股票简称`,`涨跌幅`, cast(`{key}` as float) as`{key}` FROM stock.stockdaily_vma where `{key}` > {vmaThreshold} and `日期` = "{date}" order by `{key}` DESC;'''
        results, columns = self.dbConnection.Query(sql)
        self.df = pd.DataFrame(results,columns=columns)
        self.df['涨跌幅'] = self.df['涨跌幅'].astype(float)
        self.df[key] = self.df[key].astype(float)
        results = []
        for _,row in self.df.iterrows():
            stockID = row["股票代码"]
            stockName = row["股票简称"]
            zhangDiefu = row["涨跌幅"]
            vma = row[key]
            res,data = self._Select(stockID,stockName,date,zhangDiefu,key,vma)
            if res == True:
                results.append(data)
        folderRoot= GetFuPanRoot(date) 
        datas = []
        fullPath = os.path.join(folderRoot,f"复盘摘要{date}.xlsx")
        mode='w'
        if_sheet_exists = None
        if os.path.exists(fullPath) == True:
            mode='a'
            if_sheet_exists = 'overlay'

        with pd.ExcelWriter(fullPath,engine='openpyxl',mode=mode,if_sheet_exists=if_sheet_exists) as excelWriter:
            f = CVolumnSelectorToXLSFormatter()
            f.sheetName = f"爆量复盘_{VMA}_{vmaThreshold}"
            df = pd.DataFrame()
            df.to_excel(excelWriter, sheet_name= f.sheetName,index=False)
            f.title = f"爆量复盘: {date}"
            sheet = excelWriter.sheets[f.sheetName]
            f.addTitle(sheet)     
            for result in results:
                df = result[0]
                stockID = result[1]
                stockName = result[2]
                header = result[3]
                f.AddHeader(sheet,header)
                f.AddData(excelWriter,df)
                datas.append( {"代码": stockID,"名称":stockName})
            f.formatColumnsWidth(sheet)
            
        folderRoot= GetStockFolder(date) 
        jpgDataFrame = pd.DataFrame(datas, columns=("代码","名称"))
        DataFrameToJPG(jpgDataFrame,("代码","名称"),folderRoot,f"爆量_{VMA}_{vmaThreshold}")