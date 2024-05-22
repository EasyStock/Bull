from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

class CWriteIndexMAEventToXLSX(object):
    def __init__(self,date):
        self.rows = 0
        self.date = date
        self.sheetName = f'''指数事件'''
        self.title = f'''指数事件复盘表 ({self.date})'''
    
    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 85
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,rowHight = 32):
        mergeCell = f'A{rowIndex}:E{rowIndex}'
        sheet.merge_cells(mergeCell)
        self.formatRowHeight(sheet,rowIndex,rowHight)
        for column in range(1,6):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def AddFenGeHang(self,sheet,fillColor = 'FFFFFF'):
        self.mergeRow(sheet,self.rows+1,"A","H",32)
        cell = sheet.cell(self.rows+1,1)
        cell.fill = PatternFill('solid', fgColor= fillColor)
        self.rows = self.rows + 1
    
    def WriteIndexMAEventToXLSX(self,dfToday,dfTomorrow,excelWriter):
        sheetName = f"指数均线事件复盘"
        size1 = dfToday.shape[0]
        size2 = dfTomorrow.shape[0]

        dfToday.to_excel(excelWriter, sheet_name= sheetName,index=False,startrow=1)
        dfTomorrow.to_excel(excelWriter, sheet_name= sheetName,index=False,startrow=size1+5)

        sheet = excelWriter.sheets[sheetName]
        cell = sheet.cell(1,1)
        cell.alignment = alignment
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  f"指数均线事件复盘 {self.date}"
        cell.font = Font(name='宋体', size=28, italic=False, color='000000', bold=True)
        cell.border = border
        self.mergeRow(sheet,1,40)
        startRow = 1
        endRow = startRow + size1 + size2 +5
        fontRed = Font(name='宋体', size=12, italic=False, color='FF0000', bold=True)
        fontGreen = Font(name='宋体', size=12, italic=False, color='00FF00', bold=True)
        fontblack = Font(name='宋体', size=12, italic=False, color='000000', bold=True)
        for index in range(1,endRow):
            if index >size1+1 and index < size1+5:
                continue

            if index == size1+5:
                self.mergeRow(sheet,index,40)
                cell = sheet.cell(index,1)
                cell.fill = PatternFill('solid', fgColor= 'FFFFFF')
                cell.value =  f"明日指数均线事件预测"
                cell.font = Font(name='宋体', size=28, italic=False, color='000000', bold=True)
                cell.alignment = alignment

            cell4 = sheet.cell(row = startRow + index, column = 4)
            font = fontblack
            if index >= size1 + 6:
                t = float(cell4.value)
                if t >=0:
                    font = fontRed
                else:
                    font = fontGreen

            for column in range(1,6):
                cell = sheet.cell(row = startRow + index, column = column)
                cell.border = border
                cell.alignment = alignment_left
                if column >=8 and index > 1:
                    cell.alignment = alignment_left
                cell.fill = PatternFill('solid', fgColor="FFFFFF")
                if index % 2 != 0:
                    cell.fill = PatternFill('solid', fgColor="CCEEFF")
                else:
                    cell.fill = PatternFill('solid', fgColor="FFFFFF")
                if column in (2,3):
                    cell.font = font
                else:
                    cell.font = fontblack
                
                cell.border = border

        
        self.formatColumnsWidth(sheet)

