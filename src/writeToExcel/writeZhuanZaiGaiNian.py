from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.utils import column_index_from_string
import math

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_left_top = Alignment(
         horizontal='left',
         vertical='top',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )


class CWriteZhuanZaiGaiNianToXLSX(object):
    def __init__(self,dbConnection,date):
        self.rows = 0
        self.date = date
        self.dbConnection = dbConnection
        self.sheetName = f'''转债概念'''
        self.title = f'''转债概念速查表 ({self.date})'''


    def _getZhuanZaiData(self):
        sql = f'''SELECT A.`日期`,A.`转债代码`,A.`转债名称`,A.`现价`,B.`所属概念` FROM `stock`.`kezhuanzhai_all` as A,`stock`.`stockBasicInfo` AS B where A.`正股名称`=B.`股票简称` and `日期`='{self.date}' and  A.`现价` <= 135 ;'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        reasons = []
        for _, row in df.iterrows():
            reason = row["所属概念"]
            reasons.extend(reason.split(";"))
    
        reasons = list(set(reasons))
        reasonResults = {}
        exceptGaiNian = ["融资融券","转融券标的","深股通","沪股通","富时罗素概念","富时罗素概念股","标普道琼斯A股","MSCI概念",]
        for reason in reasons:
            if reason in exceptGaiNian:
                continue
            sql = f'''SELECT A.`日期`,A.`转债代码`,A.`转债名称`,A.`现价`,B.`所属概念` FROM `stock`.`kezhuanzhai_all` as A,`stock`.`stockBasicInfo` AS B where A.`正股名称`=B.`股票简称` and `日期`='{self.date}'and B.`所属概念` like '%{reason}%' and  A.`现价` <= 135 ;'''
            result ,columns = self.dbConnection.Query(sql)
            df = pd.DataFrame(result, columns = columns)
            count = df.shape[0]
            reasonResults[reason] = (count,df['转债名称'])

        ret = sorted(reasonResults.items(), key=lambda d: d[1][0],reverse=True)
        return ret
    
    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 10
        sheet.column_dimensions['N'].width = 10
        sheet.column_dimensions['O'].width = 10
        sheet.column_dimensions['P'].width = 10
        sheet.column_dimensions['Q'].width = 10
        sheet.column_dimensions['R'].width = 10
        sheet.column_dimensions['S'].width = 10
        sheet.column_dimensions['T'].width = 15

    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeRows(self,sheet,rowstart,rowend,column1,column2,rowHight = 16):
        mergeCell = f'{column1}{rowstart}:{column2}{rowend}'
        sheet.merge_cells(mergeCell)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        if rowHight > 0:
            for row in range(rowstart,rowend+1):
                self.formatRowHeight(sheet,row,rowHight)

        for row in range(rowstart,rowend+1):
            for column in range(c1,c2+1):
                cell = sheet.cell(row = row, column = column)
                cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","T",68)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddGaiNian(self,sheet,gaiNain,data,index):
        # msg = f'''
        # {gaiNain}
        # {data[0]}
        # {data[1]}
        # '''
        # print(msg)
        # input()
        count = data[0]
        eachLine = 13
        rows =  math.ceil(count / eachLine)
        start = self.rows + 1
        end = start + rows -1

        self.mergeColumn(sheet,"A",start,end)
        sheet.column_dimensions['A'].width = 25

        self.mergeRows(sheet,start,end,"B","T",24)
        cell = sheet.cell(start,1)
        cell.alignment = alignment_center
        cell.value = gaiNain

        cell1 = sheet.cell(start,2)
        cell1.alignment = alignment_left_top
        values = ""
        for i in range(1,rows+1):
            s  = (i-1)*eachLine
            e = i * eachLine
            if e  > count:
                e = count
                values = values + "   ".join(data[1][s:count])
            else:
                values = values + "   ".join(data[1][s:e]) + "\n"
        cell1.value = values

        if index %2 == 1:
            cell.fill = PatternFill('solid', fgColor="CCEEFF")
            cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)

            cell1.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
            cell1.fill = PatternFill('solid', fgColor="CCEEFF")
        else:
            cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
            cell.fill = PatternFill('solid', fgColor="FFFFFF")

            cell1.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
            cell1.fill = PatternFill('solid', fgColor="FFFFFF")

        self.rows = self.rows + rows

    def AddFenGeHang(self,sheet,fillColor = 'FFFFFF'):
        self.mergeRow(sheet,self.rows+1,"A","T",18)
        cell = sheet.cell(self.rows+1,1)
        cell.fill = PatternFill('solid', fgColor= fillColor)
        self.rows = self.rows + 1

    def WriteZhuanZaiGainToToXLS(self,excelWriter:pd.ExcelWriter):
        tmp = pd.DataFrame()
        tmp.to_excel(excelWriter, sheet_name= self.sheetName,index=False,startrow=1)
        sheet = excelWriter.sheets[self.sheetName]
        ret = self._getZhuanZaiData()
        self.addTitle(sheet)
        for index,datas in enumerate(ret):
            gaiNian = datas[0]
            data = datas[1]
            self.AddGaiNian(sheet,gaiNian,data,index)
        self.formatColumnsWidth(sheet)