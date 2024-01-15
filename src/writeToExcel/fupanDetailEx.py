from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import column_index_from_string


# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )

class CFupanDetailEx(object):
    def __init__(self,dbConnection,tradingDays):
        self.dbConnection = dbConnection
        self.tradingDays = tradingDays
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = "详细复盘数据"
        self.title = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 10
        sheet.column_dimensions['N'].width = 10
        sheet.column_dimensions['O'].width = 10
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def _getFuPanDetail(self):
        sql = f'''SELECT * FROM stock.fupan;'''
        results, columns = self.dbConnection.Query(sql)
        
        df = pd.DataFrame(results,columns=columns)
        df.set_index(["日期",],drop=True,inplace=True)
        return df
    
    def _score(self,volumn,percentail,reversed = False):
        result = 1
        for index, value in percentail.items():
            if float(volumn) <= float(value):
                result = float(index)
                break
        
        if reversed:
            result = 1.1 - float(result)
        
        result = int(result * 10)
        return result
    
    def _WirtePart1DataToXlsx(self,df,excelWriter:pd.ExcelWriter):
        column1 = ["红盘","绿盘","两市量","实际涨停","跌停","炸板","炸板率","连板","2连板个数","3连板个数","4连板及以上个数","高度板","动能","势能"]
        df1 = pd.DataFrame(df,columns = column1)
        df1.dropna(inplace= True)
        df1['炸板率'] = df['炸板率'].apply(lambda x:x[:-1]).astype(float)
        df1['两市量'] = df['两市量'].apply(lambda x:x[:-1]).astype(float)
        for c in column1:
            df1[c] = df1[c].astype(float)
        
        rows = 10
        newDf = df1[-rows:]
        newDf.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows)
        self.rows = self.rows + rows + 1

        t = df1.quantile([0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9])
        t.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows + t.shape[0] + 4


        scoreDf = pd.DataFrame(df,columns = [])
        for c in column1:
            reversed_keys = ["绿盘","跌停","炸板","炸板率"]
            key = f'''{c}分数'''
            reversed = False
            if c in reversed_keys:
                reversed = True
            scoreDf[key] = df1.apply(lambda row: self._score(row[c],t[c],reversed), axis=1)

        rows = 10
        scoreDf[-rows:].to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows  + rows + 4
    
    def _WirtePart2DataToXlsx(self,df,excelWriter:pd.ExcelWriter):
        column2 = ["首板率","连板率","昨日首板溢价率","昨日首板晋级率","昨日2板溢价率","昨日2板晋级率","昨日3板溢价率","昨日3板晋级率","昨日4板及以上溢价率","昨日4板及以上晋级率","2进3成功率","3进高成功率"]
        df2 = pd.DataFrame(df,columns = column2)
        df2.dropna(inplace= True)
        for c in column2:
            df2[c] = df2[c].astype(float)

        rows = 10
        newDf = df2[-rows:]
        newDf.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows + rows + 4
        
        t = df2.quantile([0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9])
        t.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+ 3)
        self.rows = self.rows + t.shape[0] + 4

        scoreDf = pd.DataFrame(df,columns = [])
        for c in column2:
            reversed_keys = []
            key = f'''{c}分数'''
            reversed = False
            if c in reversed_keys:
                reversed = True
            scoreDf[key] = df2.apply(lambda row: self._score(row[c],t[c],reversed), axis=1)

        rows = 10
        scoreDf[-rows:].to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows  + rows + 4


    def _WirtePart3DataToXlsx(self,df,excelWriter:pd.ExcelWriter):
        column3 = ["涨停数量","连板数量","收-5数量","大盘红盘比","亏钱效应","首板红盘比","首板大面比","连板股的红盘比","连板比例","连板大面比","昨日连板未涨停数的绿盘比","势能EX","动能EX"]
        df3 = pd.DataFrame(df,columns = column3)
        df3.dropna(inplace= True)
        for c in column3:
            df3[c] = df3[c].astype(float)

        rows = 10
        newDf = df3[-rows:]
        newDf.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows + rows + 4


        t = df3.quantile([0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9])
        t.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows + t.shape[0] + 4

        scoreDf = pd.DataFrame(df,columns = [])
        for c in column3:
            reversed_keys = ["收-5数量","亏钱效应","首板大面比","连板大面比","昨日连板未涨停数的绿盘比",""]
            key = f'''{c}分数'''
            reversed = False
            if c in reversed_keys:
                reversed = True
            scoreDf[key] = df3.apply(lambda row: self._score(row[c],t[c],reversed), axis=1)

        rows = 10
        scoreDf[-rows:].to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows+3)
        self.rows = self.rows  + rows + 4

    def WriteFuPanDetailExToToXLS(self,excelWriter:pd.ExcelWriter):
        df = self._getFuPanDetail()
        self._WirtePart1DataToXlsx(df,excelWriter)
        self._WirtePart2DataToXlsx(df,excelWriter)
        self._WirtePart3DataToXlsx(df,excelWriter)
        sheet = excelWriter.sheets[self.sheetName]
        self.formatColumnsWidth(sheet)
        
