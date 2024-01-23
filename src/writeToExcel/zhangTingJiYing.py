from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.utils import column_index_from_string
import math

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_left_top = Alignment(
         horizontal='left',
         vertical='top',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )


class CWriteZhangTingJiYingToXLSX(object):
    def __init__(self,dbConnection,date):
        self.rows = 0
        self.date = date
        self.dbConnection = dbConnection
        self.sheetName = f'''涨停基因'''
        self.title = f'''股票连续涨停基因速查表 ({self.date})'''


    def _getZhangTingJiYingData(self):
        sql = f'''SELECT `股票代码`,`股票简称`,count(*) as `连板次数` FROM stock.stockzhangting where `连续涨停天数`>1 group by `股票代码`,`股票简称` order by `连板次数` DESC;'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        return df
    
    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20

    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeRows(self,sheet,rowstart,rowend,column1,column2,rowHight = 16):
        mergeCell = f'{column1}{rowstart}:{column2}{rowend}'
        sheet.merge_cells(mergeCell)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        if rowHight > 0:
            for row in range(rowstart,rowend+1):
                self.formatRowHeight(sheet,row,rowHight)

        for row in range(rowstart,rowend+1):
            for column in range(c1,c2+1):
                cell = sheet.cell(row = row, column = column)
                cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","J",68)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddZhangTingJiYing(self,sheet,df):
        size = df.shape[0]
        eachLine = 10
        rows =  math.ceil(size / eachLine)
        for index, row in df.iterrows():
            name = row["股票简称"]
            count = row["连板次数"]
            value = f'''{name}[{count}]'''
            row = int(index / eachLine)
            column = index % eachLine  
            cell1 = sheet.cell(row + self.rows + 1,column +1 )
            cell1.alignment = alignment_left_top
            cell1.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
            cell1.value = value
            cell1.fill = PatternFill('solid', fgColor="FFFFFF")
            if row % 2 != 0:
                cell1.fill = PatternFill('solid', fgColor="CCEEFF")

        self.rows = self.rows + rows


    def WriteZhangTingJiYingToXLS(self,excelWriter:pd.ExcelWriter):
        tmp = pd.DataFrame()
        tmp.to_excel(excelWriter, sheet_name= self.sheetName,index=False,startrow=1)
        sheet = excelWriter.sheets[self.sheetName]
        df = self._getZhangTingJiYingData()
        self.addTitle(sheet)
        self.AddZhangTingJiYing(sheet,df)
        self.formatColumnsWidth(sheet)