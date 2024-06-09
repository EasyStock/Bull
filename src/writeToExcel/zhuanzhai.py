from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import column_index_from_string
from Utility.convertDataFrameToJPG import DataFrameToJPG
from workspace import workSpaceRoot,GetZhuanZaiFolder
import os

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )

class CZhuanZaiDetail(object):
    def __init__(self,dbConnection,tradingDays):
        self.dbConnection = dbConnection
        self.tradingDays = tradingDays
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = "可转债复盘"
        self.title = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 14
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 14
        sheet.column_dimensions['H'].width = 5
        sheet.column_dimensions['I'].width = 16
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 12
        sheet.column_dimensions['L'].width = 18
        sheet.column_dimensions['M'].width = 120
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","M",140)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def ConverBigVolumnOfZhuanZhaiToJEPG(self, res):
        result = []
        for r in res:
            zhuanZaiDaiMa = r["转债代码"]
            zhuanZai_name = r["转债名称"]
            stockID = r["股票代码"]
            stockName = r["股票简称"]
            dict1 = {"代码": zhuanZaiDaiMa,"名称":zhuanZai_name}
            dict2 = {"代码": stockID,"名称":stockName}
            result.append(dict1)
            result.append(dict2)

        jpgDataFrame = pd.DataFrame(result, columns=("代码","名称"))
        folderRoot= GetZhuanZaiFolder(self.tradingDays[-1])
        DataFrameToJPG(jpgDataFrame,("代码","名称"),folderRoot,f"可转债放量")
        

    def WriteBigVolumnOfZhuanZhaiToXLSX(self,excelWriter,threshold):
        sql = f'''SELECT `转债代码` FROM stock.kezhuanzhai where `日期` = "{self.tradingDays[-1]}" and `成交额(万元)` > {threshold}'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        stockIDs = list(df["转债代码"])
        
        ret = []
        exceptGaiNian = ["融资融券","转融券标的","深股通","沪股通","富时罗素概念","富时罗素概念股","标普道琼斯A股","MSCI概念",]
        for stockID in stockIDs:
            sql1 = f'''select A.`日期`,A.`转债代码`,A.`转债名称`,A.`现价`,A.`成交额(万元)`,A.`PB`,A.`总市值（亿元)`,A.`溢价率`,A.`剩余规模`,A.`行业`,B.`股票代码`,`股票简称`,B.`所属概念` FROM `stock`.`kezhuanzhai` as A,`stock`.`stockBasicInfo` AS B where A.`正股名称`=B.`股票简称` and A.`转债代码` = "{stockID}" and A.`日期` > "{self.tradingDays[-30]}" order by A.`日期` ASC; '''
            results, columns = self.dbConnection.Query(sql1)
            df1 = pd.DataFrame(results,columns=columns)
            df1.dropna(inplace= True)

            if df1.shape[0] < 20:
                continue
            df1["成交额MA10"] = df1['成交额(万元)'].rolling(10).mean()
            df1["成交额MA20"] = df1['成交额(万元)'].rolling(20).mean()
            for e in exceptGaiNian:
                df1["所属概念"] = df1['所属概念'].str.replace(e,"")

            df1["所属概念"] = df1['所属概念'].str.replace(";+",";",regex=True).str.strip(";")

            if  df1['成交额(万元)'].iloc[-1] > 2.0* df1["成交额MA10"].iloc[-1] or df1['成交额(万元)'].iloc[-1] > 2.0* df1["成交额MA20"].iloc[-1]:
                ret.append(df1.iloc[-1].to_dict())

        self.ConverBigVolumnOfZhuanZhaiToJEPG(ret)
        newDf = pd.DataFrame(ret,columns = ["日期","转债代码","转债名称","现价","成交额(万元)","成交额MA10","成交额MA20","PB","总市值（亿元)","溢价率","剩余规模","行业","所属概念"])
        groups = newDf.groupby(["行业",])
        sortedGroups = groups["行业"].count().reset_index(name='count').sort_values(['count'], ascending=False)
        print(sortedGroups)
        resultDfs = []
        for _,row in sortedGroups.iterrows():
            hangye = row["行业"]
            df = groups.get_group(hangye).copy()
            df.sort_values(['成交额(万元)',],axis=0,ascending=False,inplace=True)
            resultDfs.append(df)
    
        newDf = pd.concat(resultDfs)
        newDf.to_excel(excelWriter, sheet_name= self.sheetName,index=False,startrow=self.rows)

        sheet = excelWriter.sheets[self.sheetName]
        startRow = self.rows  + 1
        endRow = startRow + newDf.shape[0] + 1
        for index in range(startRow,endRow):
            for column in range(1,14):
                cell = sheet.cell(row = index, column = column)
                cell.border = border
                cell.alignment = alignment_center
                cell.font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)
                if index % 2 != 0:
                    cell.fill = PatternFill('solid', fgColor="CCEEFF")
                else:
                    cell.fill = PatternFill('solid', fgColor="FFFFFF")
                
        self.rows = self.rows + newDf.shape[0]

    
    def WriteZhuanZaiInfoToExcel(self,excelWriter,threshold = 2000):
        df = pd.DataFrame()
        df.to_excel(excelWriter, sheet_name= self.sheetName,index=False)
        self.title = f"{self.tradingDays[-1]} 转债成交额大于{threshold}万元\n且大于MA10、MA20 2倍以上"
        sheet = excelWriter.sheets[self.sheetName]
        self.addTitle(sheet)
        self.WriteBigVolumnOfZhuanZhaiToXLSX(excelWriter,threshold)
        self.formatColumnsWidth(sheet)