from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import column_index_from_string
import numpy as np
from MA.MACross import CMACross

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )

class CChuangYeBanMACross(object):
    #创业板指数移动平均线相交预测表
    def __init__(self,dbConnection,tradingDays):
        self.dbConnection = dbConnection
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.tradingDays = tradingDays
        self.sheetName = "创业板均线穿插表"
        self.title = f"{self.tradingDays[-1]} 每日创业板均线穿插详细"
        self.chuangYebanMACrossDf = None
        self.changYebanCurrentValue = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12

    def formatRowsHeight(self,sheet):
        for rowIndex in range(2,11):
            sheet.row_dimensions[rowIndex].height=40
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def _getFuPanDetail(self):
        sql = f'''SELECT * FROM stock.fupan;'''
        results, columns = self.dbConnection.Query(sql)
        
        df = pd.DataFrame(results,columns=columns)
        df.set_index(["日期",],drop=True,inplace=True)
        return df
    
    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","H",60)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=28, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def CalcChuangYeBanMACross(self):
        sql = f'''SELECT * FROM stock.kaipanla_index where `StockID` = "SZ399006";'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        self.changYebanCurrentValue = df.iloc[-1]["last_px"]
        df.set_index("date",drop=True,inplace=True)
        MAs = [5,10,20,30,60,120,200]
        size = len(MAs)
        data = np.zeros((size, size))
        for i in range(size):
            for j in range(i+1,size):
                cross = CMACross(df,"last_px",MAs[i],MAs[j])
                p = cross.predict()
                print(f'''MA{MAs[i]}-MA{MAs[j]}, {p[0]:.2f},{p[1]:.2f}%''')
                data[i,j] = f'''{p[0]:.2f}'''
                data[j,i] = f'''{p[1]:.2f}'''

        columns = [f'''MA{i}''' for i in MAs]
        self.chuangYebanMACrossDf = pd.DataFrame(data,columns= columns,index=columns)
        print(self.chuangYebanMACrossDf)

    def WriteChuangYeBanMACrossToToXLS(self,excelWriter:pd.ExcelWriter):
        df = pd.DataFrame()
        df.to_excel(excelWriter, sheet_name= self.sheetName,index=False)
        sheet = excelWriter.sheets[self.sheetName]
        self.addTitle(sheet)
        self._formatChuangYeBanMACross(excelWriter)
        self.formatColumnsWidth(sheet)
        self.formatRowsHeight(sheet)

    def _formatChuangYeBanMACross(self,excelWriter:pd.ExcelWriter):
        if self.chuangYebanMACrossDf is None:
            self.CalcChuangYeBanMACross()

        self.chuangYebanMACrossDf.to_excel(excelWriter, sheet_name= self.sheetName,index=True,startrow=self.rows)
        sheet = excelWriter.sheets[self.sheetName]

        rows = 8
        # 有表头
        for r in range(self.rows+1,self.rows+rows+1):
            for c in range(1,9):
                cell = sheet.cell(row = r, column = c)
                cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
                if c>1 and (r - self.rows) == c:
                    cell.value = '-'
                
                if c>1 and (r - self.rows) > c:
                    v = float(cell.value)
                    if v >=-5 and v<=5:
                        cell.font = Font(name='宋体', size=16, italic=False, color='FF0000', bold=True)
                    cell.value = f'''{cell.value}%'''
                
                if (r - self.rows)>1 and (r - self.rows) < c:
                    v = float(cell.value)
                    if v >=self.changYebanCurrentValue*0.95 and v<=self.changYebanCurrentValue*1.05:
                        cell.font = Font(name='宋体', size=16, italic=False, color='FF0000', bold=True)

                cell.border = border
                cell.alignment = alignment_center
                
                cell.fill = PatternFill('solid', fgColor="FFFFFF")
        self.rows = self.rows + rows + 1


