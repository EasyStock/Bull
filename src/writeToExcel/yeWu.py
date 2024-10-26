from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.utils import column_index_from_string

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_left_top = Alignment(
         horizontal='left',
         vertical='top',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )


class CWriteYeWuToXLSX(object):
    def __init__(self,dbConnection,date,sql = None):
        self.rows = 0
        self.date = date
        self.sql = sql
        self.dbConnection = dbConnection
        self.sheetName = f'''可转债主营业务'''
        self.title = f'''可转债主营业务表 ({self.date})'''

    def _getZhuanZaiData(self):
        if self.sql is None:
            self.sql = f'''select  A.`转债名称`, B.`股票简称`,B.`行业`, B.`产品`,B.`范围` from (select `转债代码`, `转债名称`, `正股代码`, `正股名称` from kezhuanzhai_all where `日期` = "{self.date}")  AS A, (SELECT * FROM stock.stockyewu) AS B where A.`正股名称` = B.`股票简称`'''
        results, columns = self.dbConnection.Query(self.sql)
        df = pd.DataFrame(results,columns=columns)
        return df
    
    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 14
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 80
        sheet.column_dimensions['E'].width = 120


    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeRows(self,sheet,rowstart,rowend,column1,column2,rowHight = 16):
        mergeCell = f'{column1}{rowstart}:{column2}{rowend}'
        sheet.merge_cells(mergeCell)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        if rowHight > 0:
            for row in range(rowstart,rowend+1):
                self.formatRowHeight(sheet,row,rowHight)

        for row in range(rowstart,rowend+1):
            for column in range(c1,c2+1):
                cell = sheet.cell(row = row, column = column)
                cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","E",68)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddYeWu(self,sheet,df):
        startRow = self.rows  + 1
        endRow = startRow + df.shape[0] + 1
        for index in range(startRow,endRow):
            for column in range(1,6):
                cell = sheet.cell(row = index, column = column)
                cell.border = border
                cell.alignment = alignment_left_top
                cell.font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)

                if index == startRow:
                    cell.font = Font(name='宋体', size=14, italic=False, color='000000', bold=True)
                    cell.alignment = alignment_center

                if index % 2 != 0:
                    cell.fill = PatternFill('solid', fgColor="CCEEFF")
                else:
                    cell.fill = PatternFill('solid', fgColor="FFFFFF")
                
        self.rows = self.rows + df.shape[0]

    def WriteYeWuToToXLS(self,excelWriter:pd.ExcelWriter):
        ret = self._getZhuanZaiData()
        ret.to_excel(excelWriter, sheet_name= self.sheetName,index=False,startrow=1)
        sheet = excelWriter.sheets[self.sheetName]
        
        self.addTitle(sheet)
        self.AddYeWu(sheet,ret)
        self.formatColumnsWidth(sheet)