from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import column_index_from_string


# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left_top = Alignment(
         horizontal='left',
         vertical='top',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )

class CFupanDetail(object):
    def __init__(self,dbConnection,tradingDays):
        self.dbConnection = dbConnection
        self.tradingDays = tradingDays
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = None
        self.title = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 35
        sheet.column_dimensions['J'].width = 15
        sheet.column_dimensions['K'].width = 35
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","K",68)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddWarning(self,sheet):
        self.mergeRow(sheet,2,"A","K",68)
        cell = sheet.cell(2,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="EF4923")
        cell.value =  f'''没有超预期的票不买！！！'''
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddFenGeHang(self,sheet,fillColor = '000000'):
        self.mergeRow(sheet,self.rows+1,"A","K",32)
        cell = sheet.cell(self.rows+1,1)
        cell.fill = PatternFill('solid', fgColor= fillColor)
        self.rows = self.rows + 1


    def AddRuls(self,sheet):
        rules = [
                "1. 大盘下跌趋势不操作！！！！",									
                "2. 只做龙头不做杂毛，不浪费一颗子弹在杂毛身上，亏钱也要亏在龙头上;",									
                "3. 错过永远比错误好，对于模式内要敢于出手，模式外都要学会放弃，勿起贪念;",									
                "4. 只做周期总龙头和连板总龙头都持有和做T,降低操作频次,提高胜率;"	,								
                "5. 高潮期之后次日不开新仓，以持仓股的持有或者清仓为主;",									
                "6. 衰退期越努力越亏钱，学会空仓比什么都重要;",									
                "7. 连续三次开仓吃面，一定要休息一天，调整心态;",								
                "8.仓位管理一定要根据市场环境都赢面来确定，而不是根据自己的望断;",									
                "9.同一只股，只有第一笔仓位盈利后才能加仓;"	,						
        ]
        count = len(rules)
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)

        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  "交易原则"
        cell.font = Font(name='宋体', size=18, italic=False, color='FF0000', bold=True)
        cell.border = border
        for index,line in enumerate(rules):
            self.mergeRow(sheet,self.rows+index+1,"B","K",32)
            cell1 = sheet.cell(self.rows+index+1,2)
            cell1.value = line
            cell1.alignment = alignment_left
            cell1.fill = PatternFill('solid', fgColor="000000")
            cell1.font = Font(name='宋体', size=18, italic=False, color='FF0000', bold=True)
            cell1.border = border

        self.rows = self.rows + count

    def AddBiaozhun(self,sheet):
        qingXuBiaozhun = [
            "1. 高潮: 动能综合值=12 且 势能综合值=10 或者 连板股的红盘比 >=0.78 首板股的红盘比 >=0.78",									
            "2. 半高潮: 只有 连板股的红盘比 >=0.78"	,							
            "3. 冰点期判断 - 强势行情: 如果动能综合值 =-12 且 势能综合值 <=-2 或者 (动能综合值<=-8 且 势能综合值<=-2) 出现两次",									
            "4. 冰点期判断 - 弱势行情1: 如果动能综合值 <=-8 且 势能综合值 =-10 且首板赚钱效应和连板赚钱效应都出现过 <0.4",
            "5. 冰点期判断 - 弱势行情2: 连续两天动能综合值和势能综合值都<=-6",								
        ]
        count = len(qingXuBiaozhun)
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)

        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  "情绪判断\n标准"
        cell.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
        cell.border = border
        for index,line in enumerate(qingXuBiaozhun):
            self.mergeRow(sheet,self.rows+index+1,"B","K",32)
            cell1 = sheet.cell(self.rows+index+1,2)
            cell1.value = line
            cell1.alignment = alignment_left
            cell1.fill = PatternFill('solid', fgColor="000000")
            cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
            cell1.border = border

        self.rows = self.rows + count

    def AddYiziban(self,excelWriter:pd.ExcelWriter,sheet):
        #一字板数据
        sql = f'''SELECT `日期`,`股票代码`,`股票简称`,`连续涨停天数`,`涨停原因类别` FROM stock.yiziban where `日期` >= "{self.tradingDays[-2]}" order by `日期` ASC,`连续涨停天数` DESC;'''
        result ,columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(result, columns = columns)
        if df.empty:
            return
        count = df.shape[0] + 1
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value = f'''{self.index}.一字板数据'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        yestodayDf = df[df["日期"]== self.tradingDays[-2]]
        size = yestodayDf.shape[0]
        df.to_excel(excelWriter, sheet_name=self.sheetName, index=False,startrow=self.rows,header=True,startcol=1,merge_cells=True)
        for index in range(0,count):
            self.mergeRow(sheet,self.rows+index+1,"F","K",24)
            for c in range(1,11):
                cell1 = sheet.cell(self.rows+index+1,c+1)
                cell1.alignment = alignment_center
                if c >= 5:
                    cell1.alignment = alignment_left
                cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
                cell1.fill = PatternFill('solid', fgColor="000000")
                if index % 2 != 0:
                    cell1.fill = PatternFill('solid', fgColor="CCEEFF")
                if index > size:
                    cell1.fill = PatternFill('solid', fgColor="009DDE")
                    cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FFFFFF', bold=True)
                cell1.border = border
            
        self.index = self.index + 1
        self.rows = self.rows + count


    def AddShiChangQingxuData(self,excelWriter:pd.ExcelWriter,sheet):
        #市场情绪数据
        sql = f'''SELECT `日期`,`红盘`,`绿盘`,`两市量`,`增量`,`实际涨停`,`跌停`,`炸板`,`炸板率`,`连板` as `连板个数` FROM stock.fupan where `日期` >= "{self.tradingDays[-10]}" limit 10;'''
        result ,columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(result, columns = columns)
        if df.empty:
            return
        
        count = df.shape[0] + 2
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.市场总体\n数据'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        df.to_excel(excelWriter, sheet_name=self.sheetName, index=False,startrow=self.rows,header=True,startcol=1,merge_cells=True)
        for index in range(0,count):
            for c in range(1,11):
                cell1 = sheet.cell(self.rows+index+1,c+1)
                cell1.alignment = alignment_center
                cell1.fill = PatternFill('solid', fgColor="000000")
                if index % 2 != 0:
                    cell1.fill = PatternFill('solid', fgColor="CCEEFF")
                cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
                if index == count-2:
                    cell1.fill = PatternFill('solid', fgColor="009DDE")
                    cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FFFFFF', bold=True)
                cell1.border = border
        self.mergeRow(sheet,endRow,"B","K",24)
        cell1 = sheet.cell(endRow,2)
        cell1.alignment = alignment_left
        cell1.fill = PatternFill('solid', fgColor="000000")
        cell1.value =  f'''结论:'''
        cell1.font = Font(name='宋体', size=16, italic=False, color='FF0000', bold=True)
        cell1.border = border  
        self.index = self.index + 1
        self.rows = self.rows + count

    def AddGradeAnalysis(self,sheet):
        #各大年级段分析
        subLines = [
            "1.趋势流:",
            "2.超短流:",
            "3.投机流(可转债,新股,次新股,老妖股):",
            "4.站队/总结(a.趋势流    b.超短流   c.投机流   d.无)",
        ]
        count =len(subLines)
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.年级段分析'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        for index,line in enumerate(subLines):
            self.mergeRow(sheet,self.rows+index+1,"B","K",32)
            cell1 = sheet.cell(self.rows+index+1,2)
            cell1.value = line
            cell1.alignment = alignment_left
            cell1.fill = PatternFill('solid', fgColor="000000")
            if index % 2 != 0:
                cell1.fill = PatternFill('solid', fgColor="CCEEFF")
            cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FF0000', bold=True)
            cell1.border = border

        self.index = self.index + 1
        self.rows = self.rows + count

    def AddMacketingLove(self,sheet):
        #市场偏好
        count = 4
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.市场偏好'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border

        mergeCell = f'B{self.rows+1}:K{self.rows+3}'
        sheet.merge_cells(mergeCell)

        for index in range(0,count-1):
             self.formatRowHeight(sheet,self.rows+index+1,32)
             cell1 = sheet.cell(startRow + index,11)
             cell1.border = border
  
        cell1 = sheet.cell(startRow,2)
        cell1.alignment = alignment_left_top
        cell1.fill = PatternFill('solid', fgColor="000000")
        cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='009DDE', bold=True)
        cell1.border = border
        cell1.value = '\U000025A1短线抱团       \U000025A1趋势      \U000025A1可转债        \U000025A1新股/次新股       \U000025A1老妖股        \U000025A1二波      \U000025A1高价股        \U000025A1小盘股        \U000025A1炒数字        \U000025A1炒地图\n \
\U000025A1特殊字符(例如龙字辈)       \U000025A1一字板        \U000025A1反包      \U000025A1业绩/业绩预增        \U000025A1无'
        self.mergeRow(sheet,endRow,"B","K",24)
        cell1 = sheet.cell(endRow,2)
        cell1.alignment = alignment_left
        cell1.fill = PatternFill('solid', fgColor="000000")
        cell1.value =  f'''结论:'''
        cell1.font = Font(name='宋体', size=16, italic=False, color='FF0000', bold=True)
        cell1.border = border 

        self.index = self.index + 1
        self.rows = self.rows + count

    def AddRedianBankuai(self,sheet):
        #热点板块
        #市场情绪数据
        sql = f'''SELECT * FROM stock.rediandaily where `日期` >= "{self.tradingDays[-8]}";'''
        result ,columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(result, columns = columns)
        if df.empty:
            return
        red = InlineFont(color='009DDE',sz=self.contextFontSize)
        black = InlineFont(color='000000',sz=self.contextFontSize)
        count = 2
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value = f'''{self.index}.热点板块'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        self.mergeRow(sheet,startRow,"B","K",24)
        self.mergeRow(sheet,startRow+1,"B","K",24)

        cell1 = sheet.cell(startRow,2)
        cell1.alignment = alignment_left
        cell1.fill = PatternFill('solid', fgColor="000000")
        cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
        msg = ""
        for i in range(-6,-1):
            msg = msg + f'''{result[i][1]}({result[i][0]})    '''
        rich_string1 = CellRichText([TextBlock(black, "前5日热点: "), TextBlock(red, msg)])
        cell1.value = rich_string1

        cell2 = sheet.cell(startRow+1,2)
        cell2.alignment = alignment_left
        cell2.fill = PatternFill('solid', fgColor="000000")
        cell2.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
        rich_string2 = CellRichText([TextBlock(black, "今日热点: "), TextBlock(red, f'{result[-1][1]}')])

        rich_string2 = CellRichText([TextBlock(black, "今日热点: "), TextBlock(red, f'{result[-1][1]}({result[-1][0]})')])
        cell2.value = rich_string2

        self.index = self.index + 1
        self.rows = self.rows + count

    
    def AddQingXuZongJie1(self,excelWriter:pd.ExcelWriter,sheet):
        #市场情绪总结,1, 市场总体情绪，2，短线情绪
        sql = f'''SELECT `日期`,`10CM首板奖励率`, `10CM连板奖励率`, `20CM连板奖励率`, `20CM首板奖励率`, `连板股的红盘比`,`首板红盘比`, `动能EX` as `动能`, `势能EX` as `势能`,`备注` FROM stock.fupan where `日期` >= "{self.tradingDays[-10]}" limit 10'''
        result ,columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(result, columns = columns)
        if df.empty:
            return
        df["备注"] =df["备注"].str.replace(';', '\n')
        count = df.shape[0] + 2
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.短线情绪\n指标1'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        df.to_excel(excelWriter, sheet_name=self.sheetName, index=False,startrow=self.rows,header=True,startcol=1,merge_cells=True)
        for index in range(0,count):
            for c in range(1,11):
                cell1 = sheet.cell(self.rows+index+1,c+1)
                cell1.alignment = alignment_center
                cell1.fill = PatternFill('solid', fgColor="000000")
                if index % 2 != 0:
                    cell1.fill = PatternFill('solid', fgColor="CCEEFF")
                cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
                if index == count-2:
                    cell1.fill = PatternFill('solid', fgColor="009DDE")
                    cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FFFFFF', bold=True)
                cell1.border = border
    
        self.mergeRow(sheet,endRow,"B","K",24)
        cell1 = sheet.cell(endRow,2)
        cell1.alignment = alignment_left
        cell1.fill = PatternFill('solid', fgColor="000000")
        cell1.value =  f'''结论:'''
        cell1.font = Font(name='宋体', size=16, italic=False, color='FF0000', bold=True)
        cell1.border = border 

        self.index = self.index + 1
        self.rows = self.rows + count

    def _format1(self, text):
        texts = text.split(';')
        ret = ""
        for i in range(0, len(texts)):
            if i == 0:
                ret = texts[i] + ";"
            else:
                ret = ret + texts[i]+ ";"
                if (i+1) % 3==0:
                    ret = ret + "\n"    
        return ret
    
    def AddQingXuZongJie2(self,excelWriter:pd.ExcelWriter,sheet):
        #市场情绪总结,1, 市场总体情绪，2，短线情绪
        sql = f'''SELECT `日期`,`首板个数`, `2连板个数`, `3连板个数`, `4连板及以上个数`, `高度板`, `3连个股`,`4连及以上个股` FROM stock.fupan where `日期` >= "{self.tradingDays[-10]}" limit 10'''
        result ,columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(result, columns = columns)
        if df.empty:
            return
        df["1"] =pd.NA
        df["2"] =pd.NA
        # df["3连个股"] = df.apply(lambda row: self._format1(row['3连个股']), axis=1)
        # df["4连及以上个股"] = df.apply(lambda row: self._format1(row['4连及以上个股']), axis=1)
        newDf = pd.DataFrame(df, columns = ["日期","首板个数", "2连板个数", "3连板个数", "4连板及以上个数", "高度板", "3连个股","1","4连及以上个股","2"])
        count = df.shape[0] + 2
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.短线情绪\n指标2'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border

        newDf.to_excel(excelWriter, sheet_name=self.sheetName, index=False,startrow=self.rows,header=True,startcol=1,merge_cells=True)
        for index in range(0,count):
            self.mergeRow(sheet,startRow+index,"H","I",24)
            self.mergeRow(sheet,startRow+index,"J","K",24)
            for c in range(1,11):
                cell1 = sheet.cell(self.rows+index+1,c+1)
                cell1.alignment = alignment_center
                cell1.fill = PatternFill('solid', fgColor="000000")
                if index % 2 != 0:
                    cell1.fill = PatternFill('solid', fgColor="CCEEFF")
                cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
                if index == count-2:
                    cell1.fill = PatternFill('solid', fgColor="009DDE")
                    cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FFFFFF', bold=True)
                cell1.border = border
        
        self.mergeRow(sheet,endRow,"B","K",24)
        cell1 = sheet.cell(endRow,2)
        cell1.alignment = alignment_left
        cell1.fill = PatternFill('solid', fgColor="000000")
        cell1.value =  f'''结论:'''
        cell1.font = Font(name='宋体', size=16, italic=False, color='FF0000', bold=True)
        cell1.border = border 

        self.index = self.index + 1
        self.rows = self.rows + count

    def AddBackUpStock(self,sheet):
        #超短核心股 + 每日备选股
        self._AddNLines(sheet,"超短核心\n(每日备选股)",3)


    def AddZhouQiPanDuan(self,sheet):
        #明日周期判断
        subLines = [
            "1.超预期:",
            "2.符合预期:",
            "3.低于预期:",
            "4.总结:",
        ]
        count =len(subLines)
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.明日情绪\n判断'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        for index,line in enumerate(subLines):
            self.mergeRow(sheet,self.rows+index+1,"B","K",32)
            cell1 = sheet.cell(self.rows+index+1,2)
            cell1.value = line
            cell1.alignment = alignment_left
            cell1.fill = PatternFill('solid', fgColor="000000")
            if index % 2 != 0:
                cell1.fill = PatternFill('solid', fgColor="CCEEFF")
            cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FF0000', bold=True)
            cell1.border = border

        self.index = self.index + 1
        self.rows = self.rows + count

    def AddJinRiChaoZuo(self,sheet):
        #今日操作
        self._AddNLines(sheet,"今日操作",3)
    
    def AddMingRiJiHua(self,sheet):
        #明日计划
        self._AddNLines(sheet,"明日计划",3)
    
    def MingRiQuGuanJiHua(self,sheet):
        #明日取关计划
        self._AddNLines(sheet,"明日取关\n计划",3)

    def keZhuanZai(self,sheet):
        #可转债观察池
        self._AddNLines(sheet,"转债观察池",3)

    def _AddNLines(self,sheet,title,N):
        count = N
        startRow = self.rows+1
        endRow = startRow + count -1
        self.mergeColumn(sheet,"A",startRow,endRow)
        cell = sheet.cell(startRow,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="000000")
        cell.value =  f'''{self.index}.{title}'''
        cell.font = Font(name='宋体', size=16, italic=False, color='000000', bold=True)
        cell.border = border
        for index in range(0,count):
            self.mergeRow(sheet,self.rows+index+1,"B","K",32)
            cell1 = sheet.cell(self.rows+index+1,2)
            cell1.value = ""
            cell1.alignment = alignment_left
            cell1.fill = PatternFill('solid', fgColor="000000")
            cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
            cell1.border = border

        self.index = self.index + 1
        self.rows = self.rows + count

    def QiTa(self,sheet):
        #其他
        self._AddNLines(sheet,"其他",3)

    def WriteFuPanSummaryToXLSX(self,excelWriter:pd.ExcelWriter):
        self.sheetName = f'''股票复盘'''
        self.title = f'''每日复盘记录({self.tradingDays[-1]})'''
        tmp = pd.DataFrame()
        tmp.to_excel(excelWriter, sheet_name= self.sheetName)
        sheet = excelWriter.sheets[self.sheetName]
        self.addTitle(sheet) #标题
        self.AddWarning(sheet) #警告
        self.AddRuls(sheet) #交易规则
        self.AddFenGeHang(sheet)
        self.AddYiziban(excelWriter,sheet) #一字板数据
        self.AddFenGeHang(sheet)
        self.AddShiChangQingxuData(excelWriter,sheet)#市场情绪数据
        self.AddFenGeHang(sheet)
        self.AddGradeAnalysis(sheet)#各大年级段数据
        self.AddFenGeHang(sheet)
        self.AddMacketingLove(sheet)#市场偏好
        self.AddFenGeHang(sheet)
        self.AddRedianBankuai(sheet)#热点板块
        self.AddFenGeHang(sheet)
        self.AddQingXuZongJie1(excelWriter,sheet) #情绪总结1
        self.AddFenGeHang(sheet)
        self.AddBiaozhun(sheet) #情绪判断标准
        self.AddFenGeHang(sheet)
        self.AddQingXuZongJie2(excelWriter,sheet) #情绪总结2
        self.AddFenGeHang(sheet)
        self.AddBackUpStock(sheet) #超短核心股 + 每日备选股
        self.AddFenGeHang(sheet)
        self.AddZhouQiPanDuan(sheet)#明日周期判断
        self.AddFenGeHang(sheet)
        self.AddJinRiChaoZuo(sheet)#今日操作
        self.AddFenGeHang(sheet)
        self.AddMingRiJiHua(sheet)#明日计划
        self.AddFenGeHang(sheet)
        self.MingRiQuGuanJiHua(sheet)#明日取关计划
        self.AddFenGeHang(sheet)
        self.keZhuanZai(sheet) #可转债观察池
        self.AddFenGeHang(sheet)
        self.QiTa(sheet)#其他
        self.formatColumnsWidth(sheet)
