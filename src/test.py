from mysql.connect2DB import ConnectToDB,DataFrameToSqls_REPLACE
import pandas as pd
from thsData.fetchZhangTingFromTHS import CFetchZhangTingDataFromTHS
import numpy as np
from MA.MA import CMA
from MA.MAMgr import CMAMgr
from MA.MACross import CMACross
from Utility.convertDataFrameToJPG import DataFrameToJPG
import matplotlib.pyplot as plt
import numpy as np
from MA.MAManager import CMAManager
from DBOperating import GetTradingDateLastN
import re

def Test1_BuyTogether(dbConnection,operatorID1, operatorID2):
    sql = f'''select * from `stock`.`dragon` where (operator_ID = {operatorID1} or operator_ID = {operatorID2})  and `flag` = "B" and (date,stockID) in (select date,stockID from `stock`.`dragon` where operator_ID = {operatorID2} and `flag` = "B" and (sell = "nan" or sell = 0) and (date,stockID) in (select date,stockID from `stock`.`dragon` where operator_ID = {operatorID1} and `flag` = "B" and (sell = "nan" or sell = 0)))
'''
    print(sql)
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    print(df)

def Test1_SellTogether(dbConnection,operatorID1, operatorID2):
    sql = f'''select * from `stock`.`dragon` where (operator_ID = {operatorID1} or operator_ID = {operatorID2})  and `flag` = "S" and (date,stockID) in (select date,stockID from `stock`.`dragon` where operator_ID = {operatorID2} and `flag` = "S" and (buy = "nan" or buy = 0) and (date,stockID) in (select date,stockID from `stock`.`dragon` where operator_ID = {operatorID1} and `flag` = "S" and (buy = "nan" or buy = 0)))
'''
    print(sql)
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    print(df)



def format2To3(row):
    ban3 = row['3连板个数']
    zuoRi2Ban = row['昨日2板数']
    ratio = 0.0
    if zuoRi2Ban == 0:
        ratio = 0.0
    else:
        ratio = float(f'''{ban3/zuoRi2Ban:.2f}''')
    
    return ratio

def format3ToHigher(row):
    ban4 = row['4连板及以上个数']
    ban3_zuori = row['昨日3板数']
    ban4_zuori = row['昨日4板及以上个数']
    ratio = 0.0
    if (ban3_zuori + ban4_zuori) == 0:
        ratio = 0.0
    else:
        ratio = float(f'''{ban4/(ban3_zuori + ban4_zuori):.2f}''')
    return ratio


def DataFrameToSqls_UPDATE(datas,tableName,index_str):
    sqls = []
    for index, row in datas.iterrows():
        sql = '''UPDATE %s SET ''' %(tableName)

        for rowIndex, value in row.items():
            sql = sql + '''`%s` = '%s',''' %(rowIndex,value)
        sql = sql[:-1]
        sql = sql + ''' WHERE `%s` = '%s'; '''%(index_str,index)
        sqls.append(sql)
    return sqls

def UpdateData(dbConnection):
    sql =f'''SELECT * FROM stock.fupan;'''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)


    df["昨日2板数"] = df["2连板个数"].shift()
    df["昨日3板数"] = df["3连板个数"].shift()
    df["昨日4板及以上个数"] = df["4连板及以上个数"].shift()

    df['2进3成功率'] = df.apply(lambda row: format2To3(row), axis=1)
    df['3进高成功率'] = df.apply(lambda row: format3ToHigher(row), axis=1)
    newDf = pd.DataFrame(df,columns=['日期','2进3成功率','3进高成功率'])
    newDf.set_index(["日期",],drop=True,inplace=True)
    sqls = DataFrameToSqls_UPDATE(newDf,tableName='fupan',index_str='日期')
    #newDf.to_excel('/tmp/222.xlsx')
    for sql in sqls:
        dbConnection.Execute(sql)
        
    print(sqls[:5])


def ReadXLS():
    import openpyxl
    wb = openpyxl.load_workbook('/Volumes/Data/111.xlsx')
    print(wb.sheetnames, '\n')
    wb.add_format({'align': 'text_wrap'})
    active_sheet = wb.active
    cell = active_sheet['A1']
    #print(cell,cell.fill.bgcolor)
    print(cell,cell.value)
    print(cell,cell.font.color)

def WriteXLS():
    # import openpyxl
    # from openpyxl.cell.rich_text import TextBlock, CellRichText
    # from openpyxl.cell.text import InlineFont

    # wb = openpyxl.Workbook()
    # ws = wb.active

    # size_10 = InlineFont(sz=10,color='FF0000')
    # size_20 = InlineFont(sz=20,color='FFFF00')

    # ws["A1"] = CellRichText([TextBlock(font=size_10, text="Apple"), TextBlock(font=size_20, text="Banana")])

    # wb.save(r"/tmp/333.xlsx")

    dbConnection = ConnectToDB()
    import numpy as np
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font  # 导入字体模块
    from openpyxl.styles import PatternFill  # 导入填充模块
    from openpyxl.styles import Border,Side,Alignment,Font
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    sql = f'''SELECT * FROM stock.stockzhangting where `日期` = "2024-01-04" order by `连续涨停天数` DESC;'''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    fullPath = "/tmp/222.xlsx"
    align = Alignment(horizontal='center',vertical='center',wrap_text=True)
    fille = PatternFill('solid', fgColor= 'ffff00')  # 设置填充颜色为 橙色
    font1 = Font(name='宋体', size=28, italic=False, color="ff0000", bold=True)
    font2 = Font(name='宋体', size=28, italic=False, color="ffFFFF", bold=True)
    with pd.ExcelWriter(fullPath,engine='openpyxl',mode='w+') as excelWriter:
        pd.DataFrame().to_excel(excelWriter, sheet_name="涨停梯队", index=False,startrow=1,header=True)
        #df.to_excel(excelWriter, sheet_name="涨停梯队", index=False,startrow=1,header=True)
        ws = excelWriter.sheets['涨停梯队']
        red = InlineFont(color='FF0000')
        rich_string1 = CellRichText(['When the color ', TextBlock(red, 'red'), ' is used, you can expect ', TextBlock(red, 'danger')])

        ws.cell(1,1).value = rich_string1

    #     mergeCell = f'A{41}:H{41}'
    #     ws.merge_cells(mergeCell)
    #     ws.cell(41,1).value = rich_string1
        
        # ws.cell(1,1).alignment = align
        # ws.cell(1,1).fill = fille
        # #ws.cell(1,1).font = font1
        # ws.row_dimensions[1].height=40
        # ws.column_dimensions['A'].width = 15
        # ws.column_dimensions['B'].width = 15
        # ws.column_dimensions['C'].width = 15
        # ws.column_dimensions['D'].width = 15
        # ws.column_dimensions['E'].width = 85
        # ws.column_dimensions['F'].width = 85
        # ws.column_dimensions['G'].width = 15
        # ws.column_dimensions['H'].width = 15


def _AnalysisIndex(df,percentage = 0.5):
    lastRow = df.iloc[-1]
    newRow = lastRow.copy()
    #newRow["date"] = f'''{lastRow["date"]}_{percentage}'''
    
    newRow["last_px"] = lastRow["last_px"] * (1+percentage/100.0)
    newRow["increase_amount"] = lastRow["last_px"] * (percentage/100.0)
    newRow["increase_rate"] = f'''{percentage}%'''
    newDf = df.copy()
    newDf = newDf._append(newRow, ignore_index=False)
    return newDf

def _isAlarm(df):
    lastRow1 = df.iloc[-1]
    lastRow2 = df.iloc[-2]
    # if lastRow1["MA5_Delta"] < 0 and lastRow2["MA5_Delta"]>0:
    #     message = f'''今天是:{lastRow2.name} 明天指数涨幅小于{lastRow1["increase_rate"]},5日线将拐头向下'''
    #     print(message)
    # elif lastRow1["MA5_Delta"] > 0 and lastRow2["MA5_Delta"]<0:
    #     message = f'''今天是:{lastRow2.name} 明天指数涨幅大于{lastRow1["increase_rate"]},5日线将拐头向上'''
    #     print(message)

    if lastRow1["MA10_Delta"] < 0 and lastRow2["MA10_Delta"]>0:
        message = f'''今天是:{lastRow2.name} 明天指数涨幅小于{lastRow1["increase_rate"]},10日线将拐头向下'''
        print(message)
    elif lastRow1["MA10_Delta"] > 0 and lastRow2["MA10_Delta"]<0:
        message = f'''今天是:{lastRow2.name} 明天指数涨幅大于{lastRow1["increase_rate"]},10日线将拐头向上'''
        print(message)

    
    if lastRow1["MA5-10"] < 0 and lastRow2["MA5-10"]>0:
        message = f'''今天是:{lastRow2.name} 明天指数涨幅小于{lastRow1["increase_rate"]},5日线将 下穿 10 日线'''
        print(message)
    elif lastRow1["MA5-10"] > 0 and lastRow2["MA5-10"]<0:
        message = f'''今天是:{lastRow2.name} 明天指数涨幅大于{lastRow1["increase_rate"]},5日线将 上穿 10 日线'''
        print(message)


def AnalysisIndex():
    dbConnection = ConnectToDB()
    sql = f'''SELECT * FROM stock.kaipanla_index;'''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    chuangyeBan = df[df["StockID"] == "SZ399006"].copy()
    #chuangyeBan.reset_index(drop=True,inplace=True)
    chuangyeBan.set_index("date",drop=True,inplace=True)

    #chuangyeBan = chuangyeBan[:-50]
    mgr = CMAMgr()
    mgr.predict(chuangyeBan,"last_px",(5,10,20,30,60))
    #mgr.Cross(chuangyeBan,"last_px",5,10)

    cross = CMACross(chuangyeBan,"last_px",5,10)
    res = cross.FindAllCrossUp()
    res.to_excel("/tmp/穿.xlsx")
    # print(cross.isCrossUp(-2.46))
    # print(cross.isCrossDown(-2.46))
    # r = cross.predict()
    # print(cross.Verify(r[1]))
    #print(r)
        
    # step = list(np.linspace(-10, 10, 41))
    # for delta in step:
    #     newDf = _AnalysisIndex(chuangyeBan,delta)
    #     newDf["MA5"] = newDf["last_px"].rolling(window=5).mean()
    #     newDf["MA10"] = newDf["last_px"].rolling(window=10).mean()
    #     newDf["MA20"] = newDf["last_px"].rolling(window=20).mean()
    #     newDf["MA5_昨日"] = newDf["MA5"].shift()
    #     newDf["MA5_Delta"] = newDf["MA5"] - newDf["MA5_昨日"]

    #     newDf["MA10_昨日"] = newDf["MA10"].shift()
    #     newDf["MA10_Delta"] = newDf["MA10"] - newDf["MA10_昨日"]
    #     newDf["MA5-10"] = newDf["MA5"] - newDf["MA10"]
    #     _isAlarm(newDf)

    # # chuangyeBan["MA5_昨日"] = chuangyeBan["MA5"].shift()
    # # chuangyeBan["MA5_Delta"] = chuangyeBan["MA5"] - chuangyeBan["MA5_昨日"]

    #chuangyeBan.to_excel("/tmp/chuangyeban.xlsx")
    # results = chuangyeBan[]
    # print(chuangyeBan)
    # groups = df.groupby(["StockID",])
    # for stockID, group in groups:
    #     print(group)

def _filterZhangFu(stockID, stockName,df,threshold,days):
    result = []
    for day in range(1,days):
        key1 = f'''{day}日涨幅'''
        key2 = f'''起涨日期'''
        df[key1] = df["收盘价"].pct_change(day)
        df[key2] = df["日期"].shift(day)
        newDF = df[df[key1]>threshold]
        for _,row in newDF.iterrows():
            res = {}
            res["股票代码"] = stockID
            res["股票名称"] = stockName
            res["天数"] = day
            res[key1] = row[key1]
            res["日期"] = row["日期"]
            res[key2] = row[key2]
            result.append(res)
    return result

def FilterZhangFu():
    dbConnection = ConnectToDB()
    sql = f'''
    SELECT A.*,B.`股票简称` FROM (SELECT * FROM stock.stockdailyinfo_2023
    UNION ALL
    SELECT * FROM stock.stockdailyinfo) AS A, (SELECT * FROM stock.stockbasicinfo) AS B where A.`股票代码` = B.`股票代码`
    '''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    groups = df.groupby(["股票代码","股票简称"])
    threshold = 0.3
    days = 3
    results = []
    for (stockID,stockName),group in groups:
        df = group.reset_index()
        df.dropna()
        df["开盘价"] = df["开盘价"].astype("float")
        df["收盘价"] = df["收盘价"].astype("float")
        df["最高价"] = df["最高价"].astype("float")
        df["最低价"] = df["最低价"].astype("float")
        res = _filterZhangFu(stockID,stockName,df,threshold,days)
        results.extend(res)
    df = pd.DataFrame(results)

    df = df[df['股票名称'].str.match('[\s\S]*(ST|退|C)+?[\s\S]*') == False]
    df = df[df['股票代码'].str.match('[\s\S]*(BJ|^688)+?[\s\S]*') == False]
    df.to_excel("/tmp/区间涨幅.xlsx")


    root = "/tmp/"
    newdf = df.drop_duplicates(subset=["股票代码",],keep="first")
    DataFrameToJPG(newdf,("股票代码","股票名称"),root,f'''区间涨幅''')
    



def IndexGuaiLi():
    dbConnection = ConnectToDB()
    sql = f'''SELECT * FROM stock.kaipanla_index;'''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    chuangyeBan = df[df["StockID"] == "SZ399006"].copy()
    chuangyeBan.set_index("date",drop=True,inplace=True)
    chuangyeBan["MA5"] = chuangyeBan["last_px"].rolling(window=5).mean()
    chuangyeBan["乖离_MA5"] = (chuangyeBan["last_px"] - chuangyeBan["MA5"]) / chuangyeBan["MA5"] * 100
    chuangyeBan.to_excel("/tmp/创业板乖离.xlsx")
    t = chuangyeBan["乖离_MA5"].quantile([0.02, 0.05, 0.1, 0.5,0.9, 0.95, 0.98])
    
    print(t)
 
    # # 创建直方图
    plt.hist(chuangyeBan["乖离_MA5"], bins=100, color='green', edgecolor='black')
    
    # 设置标题和轴标签
    plt.title('Histogram of Data')
    plt.xlabel('Value')
    plt.ylabel('Frequency')
    
    # 显示图形
    plt.show()

def TestIndex():
    file = "/Users/mac/Desktop/上证指数.csv"
    df = pd.read_csv(file)
    df["MA5"] = df["last_px"].rolling(window=5).mean()
    df["MA10"] = df["last_px"].rolling(window=10).mean()
    df.to_excel("/tmp/上证指数.xlsx")
    # df["乖离"] = (df["last_px"] - df["MA5"]) / df["MA5"] * 100
    # print(df)
    # t = df["乖离"].quantile([0.005,0.01,0.025, 0.05, 0.1, 0.5,0.9, 0.95, 0.975,0.99,0.995])
    # print(t)
    # # 创建直方图
    # plt.hist(df["乖离"], bins=50, color='green', edgecolor='black')
    
    # # 设置标题和轴标签
    # plt.title('Histogram of Data')
    # plt.xlabel('Value')
    # plt.ylabel('Frequency')
    
    # # 显示图形
    # plt.show()
    cross = CMACross(df,"last_px",5,10)
    res = cross.predict()
    print(res)


def TestIndexMgr():
    dbConnection = ConnectToDB()
    sql = f'''SELECT * FROM stock.kaipanla_index;'''
    sql = f'''
    SELECT A.*,B.`股票简称` FROM (SELECT * FROM stock.stockdailyinfo_2023
    UNION ALL
    SELECT * FROM stock.stockdailyinfo) AS A, (SELECT * FROM stock.stockbasicinfo) AS B where A.`股票代码` = B.`股票代码`
    '''

    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    groups = df.groupby(["股票代码","股票简称"])
    results = []
    for (stockID,stockName),group in groups:
        df = group.reset_index()
        df.dropna()
        df.set_index("日期",drop=True,inplace=True)
        df["MA120"] = df["收盘价"].rolling(window=120).mean()
        df["MA120昨日"] = df["MA120"].shift()
        df["MA120Delta"] = df["MA120"] - df["MA120昨日"]
        df["MA120Ratio"] = (df["MA120"] - df["MA120昨日"]) / df["MA120昨日"]*10000
        ratio = df.iloc[-1]["MA120Ratio"] 
        if ratio >= -1 and ratio <= 1:
            print((stockID,stockName))
            results.append((stockID,stockName))
    newDF = pd.DataFrame(results,columns=("股票代码","股票简称"))
    newDF = newDF[newDF['股票简称'].str.match('[\s\S]*(ST|退|C)+?[\s\S]*') == False]
    newDF = newDF[newDF['股票代码'].str.match('[\s\S]*(BJ|^688)+?[\s\S]*') == False]
    newDF.to_excel("/tmp/120选股.xlsx")


    root = "/tmp/"
    # newdf = newDF.drop_duplicates(subset=["股票代码",],keep="first")
    DataFrameToJPG(newDF,("股票代码","股票简称"),root,f'''120选股''')

    #chuangyeBan = df[df["StockID"] == "SZ399006"].copy()
    # chuangyeBan.set_index("日期",drop=True,inplace=True)
    # chuangyeBan["MA120"] = chuangyeBan["收盘价"].rolling(window=120).mean()
    # chuangyeBan["MA120昨日"] = chuangyeBan["MA120"].shift()
    # chuangyeBan["MA120Delta"] = chuangyeBan["MA120"] - chuangyeBan["MA120昨日"]
    # chuangyeBan["MA120Ratio"] = (chuangyeBan["MA120"] - chuangyeBan["MA120昨日"]) / chuangyeBan["MA120昨日"]*10000
    # chuangyeBan.to_excel("/tmp/600101.xlsx")
    # mgr = CMAManager(None)
    # mgr.IndexInfo()
    

def SelectZhai():
    dbConnection = ConnectToDB()
    # from TestCode.test6 import Test
    # Test(dbConnection)
    from ZhaiSelector.ZhaiPattern1 import CZhaiPattern1
    from ZhaiSelector.ZhaiPattern2 import CZhaiPattern2

    root = "/tmp/"
    tradingDays = GetTradingDateLastN(dbConnection,300)
    params ={"startDay":tradingDays[0],"lastDay":tradingDays[-1]}
    select = CZhaiPattern1(dbConnection)
    res = select.SelectLast(params)
    print(res)

    select = CZhaiPattern2(dbConnection)
    res = select.SelectAll(params)
    res.to_excel("/tmp/140以下单日涨幅大于5%.xlsx")
    DataFrameToJPG(res,("转债代码","转债名称"),root,f'''140以下单日涨幅大于5%''')
    print(res)


def TestIndex():
    dbConnection = ConnectToDB()
    sql = f'''SELECT * FROM stock.index_dailyinfo where `指数代码` = "000001.SH";'''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    df['收盘价(点)'] = df['收盘价(点)'].astype(float)

    df["MA5"] = df["收盘价(点)"].rolling(window=5).mean()
    
    df["MA10"] = df["收盘价(点)"].rolling(window=5).mean()
    df["大于MA5"] = df["收盘价(点)"] > df["MA5"]
    df["大于MA10"] = df["收盘价(点)"] > df["MA10"]
    df.to_excel("/tmp/000001.xlsx")



def FilterZhangFuMaxZhai(start,end):
    sql = f'''
    select A.`转债代码`,A.`转债名称`,A.`现价{start}`,B.`现价{end}`,(B.`现价{end}`-A.`现价{start}`) as `delta` from (SELECT `转债代码`,`转债名称`,`现价` as `现价{start}` FROM stock.kezhuanzhai_all where `日期` = "{start}") AS A,(SELECT `转债代码` as `转债代码`,`现价` as `现价{end}` FROM stock.kezhuanzhai_all where `日期` = "{end}") AS B where A.`转债代码` = B.`转债代码` and A.`转债代码` in (SELECT `转债代码` FROM stock.kezhuanzhai where `日期` = "{end}") order by `delta` DESC  
    '''
    print(sql)
    dbConnection = ConnectToDB()
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    root = "/tmp/"
    DataFrameToJPG(df,("转债代码","转债名称"),root,f'''区间涨幅排名{end}''')
    df.to_excel(f'''/tmp/区间涨幅排名{start} - {end}.xlsx''')


def HongShanBing():
    dbConnection = ConnectToDB()
    sql = f'''    SELECT A.*,B.`股票简称` FROM (SELECT * FROM stock.stockdailyinfo) AS A, (SELECT * FROM stock.stockbasicinfo) AS B where A.`股票代码` = B.`股票代码`'''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    groups = df.groupby(["股票代码",])

    result = []
    for stockID, group in groups:
        if group.shape[0] <10:
            continue

        group["开盘价"] = group["开盘价"].astype("float")
        group["收盘价"] = group["收盘价"].astype("float")
        group["最低价"] = group["最低价"].astype("float")
        group["成交量"] = group["成交量"].astype("float")
        group["昨日收盘价"] = group["收盘价"].shift()
        group["涨跌额"] = group["收盘价"] - group["昨日收盘价"]

        group["昨日成交量"] = group["成交量"].shift()
        group["涨跌量"] = group["成交量"] - group["昨日成交量"]

        if group.iloc[-1]["涨跌额"] <=0 or group.iloc[-2]["涨跌额"] <=0:
            continue

        if group.iloc[-1]["涨跌量"] <=0 or group.iloc[-2]["涨跌量"] <=0:
            continue

        if group.iloc[-1]["收盘价"] <=group.iloc[-1]["开盘价"] or group.iloc[-2]["收盘价"] <=group.iloc[-2]["开盘价"] or group.iloc[-3]["收盘价"] <=group.iloc[-3]["开盘价"]:
            continue

        if group.iloc[-1]["涨跌额"] <= group.iloc[-2]["涨跌额"]:
            continue

        if group.iloc[-1]["涨跌量"] <= group.iloc[-2]["涨跌量"]:
            continue


        if group.iloc[-2]["涨跌额"] <= group.iloc[-3]["涨跌额"]:
            continue

        if group.iloc[-2]["涨跌量"] <= group.iloc[-3]["涨跌量"]:
            continue
        
        group[-15:].to_excel(f'''/tmp/红三兵{stockID[0]}.xlsx''')
        result.append({"股票代码":stockID[0],"股票名称":group.iloc[-1]["股票简称"]})
        print(f"================================{stockID[0]} Done!!!!================================")

    newdf = pd.DataFrame(result)
    DataFrameToJPG(newdf,("股票代码","股票名称"),"/tmp/",f'''红三兵''')




def MATest():
    dbConnection = ConnectToDB()
    sql = f'''
    SELECT * FROM stock.stockdailyinfo_2021
    UNION ALL
    SELECT * FROM stock.stockdailyinfo_2022
    UNION ALL
    SELECT * FROM stock.stockdailyinfo_2023
    UNION ALL
    SELECT * FROM stock.stockdailyinfo_2024
    UNION ALL
    SELECT * FROM stock.stockdailyinfo
    '''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    df.drop_duplicates()
    groups = df.groupby(["股票代码",])

    result = []
    for stockID, group in groups:
        if group.shape[0] <10:
            continue

        group["开盘价"] = group["开盘价"].astype("float").round(3)
        group["收盘价"] = group["收盘价"].astype("float").round(3)
        group["最低价"] = group["最低价"].astype("float").round(3)
        group["成交量"] = group["成交量"].astype("float").round(3)

        group["MA5"] = group["收盘价"].rolling(window=5).mean().round(3)
        group["收盘价-MA5"] = (group["收盘价"] - group["MA5"]).round(3)

        group["MA10"] = group["收盘价"].rolling(window=10).mean().round(3)
        group["收盘价-MA10"] = (group["收盘价"] - group["MA10"]).round(3)

        group["MA20"] = group["收盘价"].rolling(window=20).mean().round(3)
        group["收盘价-MA20"] = (group["收盘价"] - group["MA20"]).round(3)

        group["MA30"] = group["收盘价"].rolling(window=30).mean().round(3)
        group["收盘价-MA30"] = (group["收盘价"] - group["MA30"]).round(3)

        group["MA60"] = group["收盘价"].rolling(window=60).mean().round(3)
        group["收盘价-MA60"] = (group["收盘价"] - group["MA60"]).round(3)

        group["MA120"] = group["收盘价"].rolling(window=120).mean().round(3)
        group["收盘价-MA120"] = (group["收盘价"] - group["MA120"]).round(3)

        group["MA250"] = group["收盘价"].rolling(window=250).mean().round(3)
        group["收盘价-MA250"] = (group["收盘价"] - group["MA250"]).round(3)
        
        columns = ["日期","股票代码","收盘价","MA5","MA10","MA20","MA30","MA60","MA120","MA250","收盘价-MA5","收盘价-MA10","收盘价-MA20","收盘价-MA30","收盘价-MA60","收盘价-MA120","收盘价-MA250"]
        newdf = pd.DataFrame(data=group, columns=columns)
        sqls = DataFrameToSqls_REPLACE(datas=newdf,tableName="stock_daily_info_mas")
        step = 300
        groupedSql = [" ".join(sqls[i:i+step]) for i in range(0,len(sqls),step)]
        for sql in groupedSql:
            dbConnection.Execute(sql)
        
        print("stockID: " + str(stockID))


    print("================================end================================")


def MATest_statistics():
    dbConnection = ConnectToDB()
    sql = f'''
    SELECT * FROM stock.stock_daily_info_mas;
    '''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    df.drop_duplicates()

    groups = df.groupby(["日期",])
    result = []
    for date, group in groups:
        if group.shape[0] <10:
            continue
        
        print(f'''================================开始处理{date[0]}================================''')
        group["收盘价"] = group["收盘价"].astype("float").round(3)

        group["MA5"] = group["MA5"].astype("float").round(3)
        group["MA10"] = group["MA10"].astype("float").round(3)
        group["MA20"] = group["MA10"].astype("float").round(3)
        group["MA30"] = group["MA10"].astype("float").round(3)
        group["MA60"] = group["MA10"].astype("float").round(3)
        group["MA120"] = group["MA10"].astype("float").round(3)
        group["MA250"] = group["MA10"].astype("float").round(3)

        group["收盘价-MA5"] = group["收盘价-MA5"].astype("float").round(3)
        group["收盘价-MA10"] = group["收盘价-MA10"].astype("float").round(3)
        group["收盘价-MA20"] = group["收盘价-MA20"].astype("float").round(3)
        group["收盘价-MA30"] = group["收盘价-MA30"].astype("float").round(3)
        group["收盘价-MA60"] = group["收盘价-MA60"].astype("float").round(3)
        group["收盘价-MA120"] = group["收盘价-MA120"].astype("float").round(3)
        group["收盘价-MA250"] = group["收盘价-MA250"].astype("float").round(3)

        group["MA5乖离率"] = group["收盘价-MA5"]/group["收盘价"]*100
        group["MA10乖离率"] = group["收盘价-MA10"]/group["收盘价"]*100
        group["MA20乖离率"] = group["收盘价-MA20"]/group["收盘价"]*100
        group["MA30乖离率"] = group["收盘价-MA30"]/group["收盘价"]*100
        group["MA60乖离率"] = group["收盘价-MA60"]/group["收盘价"]*100
        group["MA120乖离率"] = group["收盘价-MA120"]/group["收盘价"]*100
        group["MA250乖离率"] = group["收盘价-MA250"]/group["收盘价"]*100


        shenzheng = group[group['股票代码'].str.match('^00.*') == True].copy()
        shenzheng["股票市场"] = "中小板"

        shangzheng = group[group['股票代码'].str.match('^60.*') == True].copy()
        shangzheng["股票市场"] = "主板"

        chuangyeBan = group[group['股票代码'].str.match('^30.*') == True].copy()
        chuangyeBan["股票市场"] = "创业板"

        kezhuangBan = group[group['股票代码'].str.match('^68.*') == True].copy()
        kezhuangBan["股票市场"] = "科创板"

        beijiaosuo = group[group['股票代码'].str.match('.*BJ') == True].copy()
        beijiaosuo["股票市场"] = "北交所"

        others = group[group['股票代码'].str.match('^00.*|^60.*|^30.*|^68.*|.*BJ') == False]

        if others.shape[0] > 0:
            raise Exception("unknown 类型")
        

        allDF = [shenzheng,shangzheng,chuangyeBan,kezhuangBan,beijiaosuo]
        for shiChangDF in allDF:
            r = {}
            r["日期"] = date[0]
            if shiChangDF.empty:
                continue

            r["股票市场"] = shiChangDF.iloc[0]["股票市场"]
            count = shiChangDF.shape[0]
            r[f'''总数'''] = count
            
            
            columns = ["MA5","MA10","MA20","MA30","MA60","MA120","MA250"]
            for column in columns:
                basicColumns = ["日期","股票代码","收盘价"]
                basicColumns.append(column)
                key1 = f'''收盘价-{column}'''
                key2 = f'''{column}乖离率'''
                basicColumns.append(key1)
                basicColumns.append(key2)
                newDF = pd.DataFrame(data=shiChangDF, columns=basicColumns)
                newDF.dropna(inplace=True)
                if newDF.empty:
                    r[f'''大于{column}数量'''] = ""
                    r[f'''大于{column}百分比'''] = ""
                    r[f'''{column}平均乖离率'''] = ""
                    continue

                count1 = newDF[newDF[key1] >= 0].shape[0]
                r[f'''大于{column}数量'''] = count1
                r[f'''大于{column}百分比'''] = f'''{count1/count*100:.2f}'''
                r[f'''{column}平均乖离率'''] = f'''{newDF[key2].mean():.2f}'''
            
            result.append(r)
    
            
    resultDf = pd.DataFrame(data=result)
    print(resultDf)
    sqls = DataFrameToSqls_REPLACE(datas=resultDf,tableName="stock_daily_info_mas_stastics")
    step = 300
    groupedSql = [" ".join(sqls[i:i+step]) for i in range(0,len(sqls),step)]
    for sql in groupedSql:
        dbConnection.Execute(sql)


def N_statistics():
    dbConnection = ConnectToDB()
    sql = f'''
    SELECT * FROM stock.stockzhangting where `日期` >= "2025-04-01"
    '''
    results, columns = dbConnection.Query(sql)
    df = pd.DataFrame(results,columns=columns)
    print(df)
    groups = df.groupby(["股票代码",])
    result = []
    for stockID, group in groups:
        if group.shape[0] >= 2:
            continue

        result.append([group.iloc[0]["股票代码"],group.iloc[0]["股票简称"]])
    
    newDF = pd.DataFrame(data = result,columns=("股票代码","股票简称"))
    DataFrameToJPG(newDF,("股票代码","股票简称"),"/tmp/",f'''N型备选''')
    print(result)

    # To DO ：
    # 先涨停，然后三根阴线

if __name__ == "__main__":
    # dbConnection = ConnectToDB()
    # # Test1_BuyTogether(dbConnection,10028451,10656871)
    # # Test1_SellTogether(dbConnection,10028451,10656871)
    # #UpdateData(dbConnection)
    # #WriteXLS()
    # # WriteXLS()
    # #AnalysisIndex()
    # #FilterZhangFu()
    # #TestIndex()
    # tradingDays = GetTradingDateLastN(dbConnection,3)
    #FilterZhangFuMaxZhai("2024-09-20","2024-10-08")
    # import textwrap

    # text = '危楼高百尺，手可摘星辰。不敢高声语，恐惊天上人。'
    # width = 12
    # wrapped_text = textwrap.wrap(text, width)
    # for line in wrapped_text:
    #     print(line)
    HongShanBing()
    # MATest()
    # MATest_statistics()
    #N_statistics()