import os
import pandas as pd
import numpy as np
import logging
from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd
from openpyxl.utils import column_index_from_string
from Utility.convertDataFrameToJPG import DataFrameToJPG
from workspace import workSpaceRoot,GetFuPanRoot
import os

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )

logger = logging.getLogger()

class CPercentileBanKuai(object):
    def __init__(self,dbConnection):
        self.dbConnection = dbConnection

    def _formatVolumn(self,volumn:str):
        newVolumn = float(volumn[:-1])
        danwei = volumn[-1]
        ret = volumn
        if danwei == "万":
            ret = newVolumn*10000
        if danwei == "亿":
            ret = newVolumn*100000000
        return ret
    
    def _readExceptData(self):
        folder = os.path.abspath(os.path.dirname(__file__))
        banKuaifile = os.path.join(folder,"bankuai_except.xlsx")
        df_except = pd.read_excel(banKuaifile)
        return df_except

    def _DataFrameToSqls_REPLACE_INTO(self,datas,tableName):
        sqls = []
        for _, row in datas.iterrows():
            index_str = '''`,`'''.join(row.index)
            value_str = '''","'''.join(str(x) for x in row.values)
            sql = '''REPLACE INTO {0} (`{1}`) VALUES ("{2}");'''.format(tableName,index_str,value_str)
            sqls.append(sql)
        return sqls
    
    def _DataFrameToSqls_UPDATE(self,datas,tableName,index_str):
        sqls = []
        for index, row in datas.iterrows():
            sql = '''UPDATE %s SET ''' %(tableName)

            for rowIndex, value in row.items():
                sql = sql + '''`%s` = '%s',''' %(rowIndex,value)
            sql = sql[:-1]
            
            if isinstance(index_str,str):
                sql = sql + ''' WHERE `%s` = '%s'; '''%(index_str,index)
            elif isinstance(index_str, (list, tuple)):
                ziped = dict(zip(index_str,index))
                indexes = ""
                for key in ziped:
                    indexes = indexes + f'''(`{key}` = '{ziped[key]}') AND '''
                
                indexes = indexes[:-4] # remove last "AND "
                sql = sql + ''' WHERE %s ; '''%(indexes)
            sqls.append(sql)
        return sqls
    

    def PercentileBanKuai_History(self):
        sql = f'''SELECT * FROM stock.bankuai_index_dailyinfo;'''
        result1,columns1 = self.dbConnection.Query(sql)
        df_all=pd.DataFrame(result1,columns=columns1)

        df_except = self._readExceptData()
        newDF = df_all[~df_all.板块代码.isin(df_except.板块代码)]

        groups = newDF.groupby(["板块代码","板块名称"])

        step = list(np.linspace(0.01, 1, 100))
        for stockID, group in groups:
            if group.shape[0] < 10:
                continue
            group["成交量(股)"] = group.apply(lambda row:self._formatVolumn(row["成交量(股)"]),axis=1)
            group["成交额(元)"] = group.apply(lambda row:self._formatVolumn(row["成交额(元)"]),axis=1)
            group["流通市值(元)"] = group.apply(lambda row:self._formatVolumn(row["流通市值(元)"]),axis=1)
            group["总市值(元)"] = group.apply(lambda row:self._formatVolumn(row["总市值(元)"]),axis=1)
            group.reset_index(inplace=True)
            columns = ["开盘价(点)","收盘价(点)","最高价(点)","最低价(点)","成交量(股)","成交额(元)","涨跌幅(%)","量比","换手率(%)","上涨家数(家)","下跌家数(家)","流通市值(元)","总市值(元)"]
            newDf = pd.DataFrame(group, columns=columns)
            for c in columns:
                newDf[c] = newDf[c].astype(float)
                
            percentile = newDf.quantile(step)
            percentile["百分位数"] = percentile.index
            percentile["板块代码"] = stockID[0]
            percentile["板块名称"] = stockID[1]

            for c in columns:
                percentile[c] = percentile.apply(lambda row: '{:.2f}'.format(row[c]), axis=1)

            percentile["百分位数"] = percentile.apply(lambda row: '{:.02f}'.format(row["百分位数"]), axis=1)
            percentile.reset_index(inplace=True,drop=True)

            #t.set_index(["板块代码","百分位数"],drop=True,inplace=True)
            sqls = self._DataFrameToSqls_REPLACE_INTO(percentile,"percentile_bankuai")
            sql2 = " ".join(sqls)
            self.dbConnection.Execute(sql2)

    
    def _GetBankuaiPercentileHistory(self):
        sql = f'''SELECT * FROM stock.percentile_bankuai;'''
        result1,columns1 = self.dbConnection.Query(sql)
        df_all=pd.DataFrame(result1,columns=columns1)
        return df_all
    
    def _score(self,volumn,percentail,reversed = False):
        result = 1
        for index, value in percentail.items():
            if float(volumn) <= float(value):
                result = float(index)
                break
        
        if reversed:
            result = 1.0 - float(result)
        
        result = int(result * 100)
        return result
    
    def _processOneDay(self,day,df:pd.DataFrame):
        logger.info(f'''=========开始处理: {day}=========''')
        columns = ["日期","板块代码","板块名称","涨跌幅(%)",]
        newDf = pd.DataFrame(df,columns = columns)
        newDf.set_index(["日期","板块代码","板块名称"],drop=True,inplace=True)
        newDf["涨跌幅(%)"] = newDf["涨跌幅(%)"].astype(float)
        step = list(np.linspace(0.01, 1, 100))
        percentile = newDf.quantile(step)
        newDf["涨跌幅相对分数"] =newDf.apply(lambda row: self._score(row["涨跌幅(%)"],percentile["涨跌幅(%)"],False), axis=1)
        newDf.reset_index(["板块名称",],drop=True,inplace=True)
        newDf.drop(["涨跌幅(%)",],axis=1,inplace=True)
        sqls = self._DataFrameToSqls_UPDATE(newDf,"bankuai_index_score_daily",["日期","板块代码"])
        sql = " ".join(sqls)
        return sql

    def _prcessLastNDays(self,newDF:pd.DataFrame,lastNDays):
        sqls = []
        groupByDate = newDF.groupby(["日期"])
        for day in lastNDays:
            try:
                df = groupByDate.get_group(day)
            except:
                continue
            sql = self._processOneDay(day,df)
            sqls.append(sql)
        
        return sqls


    def _processHistoryData(self,percentileDf:pd.DataFrame,newDF:pd.DataFrame,LastNDays):
        groupByStockID = newDF.groupby(["板块代码","板块名称"])
        sqls = []
        for stock, group in groupByStockID:
            try:
                percentile = percentileDf.get_group(stock)
            except:
                continue
            percentile.set_index(["百分位数",],drop=True,inplace=True)
            group["成交量(股)"] = group.apply(lambda row:self._formatVolumn(row["成交量(股)"]),axis=1)
            group["成交额(元)"] = group.apply(lambda row:self._formatVolumn(row["成交额(元)"]),axis=1)
            group["流通市值(元)"] = group.apply(lambda row:self._formatVolumn(row["流通市值(元)"]),axis=1)
            group["总市值(元)"] = group.apply(lambda row:self._formatVolumn(row["总市值(元)"]),axis=1)
            df = pd.DataFrame(group)
            df.reset_index(drop=True,inplace=True)
            size = len(LastNDays)
            df = df[-size:]
            df["开盘价分数"] =df.apply(lambda row: self._score(row["开盘价(点)"],percentile["开盘价(点)"],False), axis=1)
            df["收盘价分数"] =df.apply(lambda row: self._score(row["收盘价(点)"],percentile["收盘价(点)"],False), axis=1)
            df["最高价分数"] =df.apply(lambda row: self._score(row["最高价(点)"],percentile["最高价(点)"],False), axis=1)
            df["最低价分数"] =df.apply(lambda row: self._score(row["最低价(点)"],percentile["最低价(点)"],False), axis=1)
            df["成交量分数"] =df.apply(lambda row: self._score(row["成交量(股)"],percentile["成交量(股)"],False), axis=1)
            df["成交额分数"] =df.apply(lambda row: self._score(row["成交额(元)"],percentile["成交额(元)"],False), axis=1)
            df["涨跌幅自身分数"] =df.apply(lambda row: self._score(row["涨跌幅(%)"],percentile["涨跌幅(%)"],False), axis=1)
            df["量比分数"] =df.apply(lambda row: self._score(row["量比"],percentile["量比"],False), axis=1)
            df["换手率分数"] =df.apply(lambda row: self._score(row["换手率(%)"],percentile["换手率(%)"],False), axis=1)
            df["上涨家数分数"] =df.apply(lambda row: self._score(row["上涨家数(家)"],percentile["上涨家数(家)"],False), axis=1)
            df["下跌家数分数"] =df.apply(lambda row: self._score(row["下跌家数(家)"],percentile["下跌家数(家)"],True), axis=1)
            df["流通市值分数"] =df.apply(lambda row: self._score(row["流通市值(元)"],percentile["流通市值(元)"],False), axis=1)
            df["总市值分数"] =df.apply(lambda row: self._score(row["总市值(元)"],percentile["总市值(元)"],False), axis=1)
            df.drop(["开盘价(点)","收盘价(点)","最高价(点)","最低价(点)","成交量(股)","成交额(元)","涨跌幅(%)","量比","换手率(%)","上涨家数(家)","下跌家数(家)","流通市值(元)","总市值(元)","顶底分型"],axis=1,inplace=True)
            groupSqls = self._DataFrameToSqls_REPLACE_INTO(df,"bankuai_index_score_daily")
            sql = " ".join(groupSqls)
            sqls.append(sql)
        return sqls

    def PercentileBankuai_LastNDays(self,LastNDays:list):
        banKuaiHistory = self._GetBankuaiPercentileHistory()
        percentiles = banKuaiHistory.groupby(["板块代码","板块名称"])

        sql = f'''SELECT * FROM stock.bankuai_index_dailyinfo;'''
        result1,columns1 = self.dbConnection.Query(sql)
        df_all=pd.DataFrame(result1,columns=columns1)
        df_except = self._readExceptData()
        newDF = df_all[~df_all.板块代码.isin(df_except.板块代码)]

        sqls = self._processHistoryData(percentiles,newDF,LastNDays)
        sqls1 = self._prcessLastNDays(newDF,LastNDays)
        sqls.extend(sqls1)
        
        for sql in sqls:
            self.dbConnection.Execute(sql)
        


class CBanKuaiSelectorToXLSFormatter(object):
    def __init__(self):
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = "板块复盘"
        self.title = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['E'].width = 14
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 14
        sheet.column_dimensions['H'].width = 14
        sheet.column_dimensions['I'].width = 14
        sheet.column_dimensions['J'].width = 14
        sheet.column_dimensions['K'].width = 14
        sheet.column_dimensions['L'].width = 14
        sheet.column_dimensions['M'].width = 14
        sheet.column_dimensions['N'].width = 16
        sheet.column_dimensions['O'].width = 16
        sheet.column_dimensions['P'].width = 16
        sheet.column_dimensions['Q'].width = 14
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","Q",70)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 2

    def AddHeader(self,sheet,header):
        self.mergeRow(sheet,self.rows+1,"A","Q",30)
        cell = sheet.cell(self.rows+1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  header
        cell.font = Font(name='宋体', size=18, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddData(self,excelWriter,df:pd.DataFrame):
        df.to_excel(excelWriter, sheet_name= self.sheetName,index=False,startrow=self.rows)
        sheet = excelWriter.sheets[self.sheetName]
        startRow = self.rows  + 1
        endRow = startRow + df.shape[0] + 1
        for index in range(startRow,endRow):
            for column in range(1,18):
                cell = sheet.cell(row = index, column = column)
                cell.border = border
                cell.alignment = alignment_center
                cell.font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)
                if index % 2 != 0:
                    cell.fill = PatternFill('solid', fgColor="CCEEFF")
                else:
                    cell.fill = PatternFill('solid', fgColor="FFFFFF")
                
        self.rows = self.rows + df.shape[0] + 3



class CBanKuaiSelect(object):
    def __init__(self,dbConnection):
        self.dbConnection = dbConnection

    def BanKuaiSelect(self,excelWriter,tradingDays,limit = 50):
        sqls = []
        for day in tradingDays:
            sql = f'''(SELECT * FROM stock.bankuai_index_score_daily where `日期`= "{day}" order by `涨跌幅相对分数` DESC limit {limit})'''
            sqls.append(sql)
        
        newSql = "\n UNION ALL \n".join(sqls)
        result1,columns1 = self.dbConnection.Query(newSql)
        df_all=pd.DataFrame(result1,columns=columns1)
        result = {}
        for _,row in df_all.iterrows():
            stockID = row["板块代码"]
            stockName = row["板块名称"]
            date = row["日期"]
            zhangdiefu = row["涨跌幅相对分数"]
            if stockID not in result:
                result[stockID] = [stockID,stockName,date,zhangdiefu,zhangdiefu,1]
            else:
                oldValue = result[stockID]
                date1 = f'''{oldValue[2]} {date}'''
                zongfen =  oldValue[3] + zhangdiefu
                count = oldValue[5] + 1
                avg = f'''{zongfen/count:.2f}'''
                result[stockID] = [stockID,stockName,date1,zongfen,avg,count]
        
        t = [result[x] for x in result]
        df = pd.DataFrame(t,columns = ["板块代码","板块名称","日期","总分","平均分","个数"])
        df = df[df["个数"] >= 2]
        df.sort_values(['个数',"平均分"],axis=0,ascending=[False,True],inplace=True)
        df.reset_index(inplace=True,drop=True)
        folder = GetFuPanRoot(tradingDays[-1])
        DataFrameToJPG(df,["板块代码","板块名称"],folder,"板块复盘")
        self.WriteToXLSX(excelWriter,df_all,df)


    def WriteToXLSX(self,excelWriter,df_all,resultdf):
        f = CBanKuaiSelectorToXLSFormatter()
        f.sheetName = f"板块复盘"
        df = pd.DataFrame()
        df.to_excel(excelWriter, sheet_name= f.sheetName,index=False)
        f.title = f"板块复盘"
        sheet = excelWriter.sheets[f.sheetName]
        f.addTitle(sheet) 
        for _,row in resultdf.iterrows():
            stockID = row["板块代码"]
            df = df_all[df_all["板块代码"] == stockID]
            df.reset_index(inplace=True,drop = True)
            header = f'''板块代码: {row["板块代码"]}  板块名称: {row["板块名称"]}   总分: {row["总分"]}     平均分: {row["平均分"]}     个数: {row["个数"]}'''
            f.AddHeader(sheet,header)
            f.AddData(excelWriter,df)
        f.formatColumnsWidth(sheet)


def _GetAllBanKuaiInfo(dbconnection):
    sql = f'''SELECT `板块代码`,`板块名称` FROM stock.bankuai_index_dailyinfo group by `板块代码`,`板块名称`;'''
    result1,columns1 = dbconnection.Query(sql)
    df=pd.DataFrame(result1,columns=columns1)
    return df

def _GetAllValueableBanKuaiData(dbconnection,otherDf):
    folder = os.path.abspath(os.path.dirname(__file__))
    banKuaifile = os.path.join(folder,"bankuai_except.xlsx")
    df_except = pd.read_excel(banKuaifile)
    res = otherDf[~otherDf.板块代码.isin(df_except.板块代码)]
    return res

def _GetAllBanKuaiData(dbConnection):
    sql = f'''SELECT * FROM stock.bankuai_index_dailyinfo;'''
    result1,columns1 = dbConnection.Query(sql)
    df=pd.DataFrame(result1,columns=columns1)
    return df


def GetAllValueableBanKuaiData(dbConnection):
    df = _GetAllBanKuaiData(dbConnection)
    res = _GetAllValueableBanKuaiData(dbConnection,df)
    print(res)

def GetAllBasicBanKuaiData(dbConnection):
    df = _GetAllBanKuaiInfo(dbConnection)
    res = _GetAllValueableBanKuaiData(dbConnection,df)
    res.to_excel("/tmp/_bankuai_except_.xlsx",index=False)
    return res

