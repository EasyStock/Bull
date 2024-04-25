import os
import pandas as pd
import numpy as np
from workspace import workSpaceRoot,WorkSpaceFont,GetZhuanZaiFolder
from Utility.convertDataFrameToJPG import DataFrameToJPG

from workspace import workSpaceRoot,WorkSpaceFont,GetZhuanZaiFolder,GetFuPanRoot
from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
from openpyxl.utils import column_index_from_string

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )


class CWriteZaiScoreToXlsx(object):
    def __init__(self):
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = "可转债分数快速增加"
        self.title = "可转债分数快速增加速查表"

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 18
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 18
        sheet.column_dimensions['H'].width = 18
        sheet.column_dimensions['I'].width = 18
        sheet.column_dimensions['J'].width = 18
        sheet.column_dimensions['K'].width = 18
        sheet.column_dimensions['L'].width = 18
        sheet.column_dimensions['M'].width = 18
        sheet.column_dimensions['N'].width = 18
        sheet.column_dimensions['O'].width = 18
        sheet.column_dimensions['P'].width = 18
        sheet.column_dimensions['Q'].width = 18
        sheet.column_dimensions['R'].width = 18
        sheet.column_dimensions['S'].width = 18
        sheet.column_dimensions['T'].width = 18

    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","T",140)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddFirstLine(self,sheet):
        self.mergeRow(sheet,self.rows+1,"A","T",100)
        cell = sheet.cell(self.rows+1,1)
        cell.alignment = alignment_left
        cell.fill = PatternFill('solid', fgColor="FFFFFF")
        cell.value = f'''说明:\n1.红色部分说明今日分数超过了分位BOLL上轨.\n2.紫色部分说明数据等于最大值.'''
        cell.font = Font(name='宋体', size= self.contextFontSize, italic=False, color='FF0000', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddTail(self,sheet):
        self.mergeRow(sheet,self.rows+1,"A","T",32)
        cell = sheet.cell(self.rows+1,1)
        cell.alignment = alignment_left
        cell.fill = PatternFill('solid', fgColor="FFFFFF")
        cell.value =  f'''风险提示: 本内容仅信息分享,不构成投资建议,若以此作为买卖依据,后果自负。市场有风险,投资需谨慎！'''
        cell.font = Font(name='宋体', size= 10, italic=False, color='FF0000', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def WriteScoreToXLSX(self, scoreDF:pd.DataFrame,excelWriter:pd.ExcelWriter):
        rows = scoreDF.shape[0] + 1
        columns = scoreDF.shape[1]+1
        df = pd.DataFrame()
        df.to_excel(excelWriter, sheet_name= self.sheetName,index=False)
        sheet = excelWriter.sheets[self.sheetName]
        self.addTitle(sheet)
        self.AddFirstLine(sheet)
        scoreDF.index = scoreDF.index +1
        scoreDF.to_excel(excelWriter, sheet_name= self.sheetName,index=True,index_label = "序号",startrow=self.rows,header=True)
        for index in range(0,rows):
            scoreToday = scoreDF.iloc[index -1]["今日分数"]
            max = scoreDF.iloc[index -1]["最高分"]
            score98 = scoreDF.iloc[index -1]["BOLL上轨"]
            
            fontMax =  None
            fontBoll =  None
            if scoreToday >= max:
                fontMax = Font(name='宋体', size=self.contextFontSize, italic=False, color='CC33FF', bold=True)
            elif scoreToday >= score98:
                fontBoll = Font(name='宋体', size=self.contextFontSize, italic=False, color='FF6666', bold=True)
        
            for c in range(1,columns+1):
                cell1 = sheet.cell(self.rows+index+1,c)
                cell1.alignment = alignment_center
                cell1.fill = PatternFill('solid', fgColor="FFFFFF")
                if index % 2 != 0:
                    cell1.fill = PatternFill('solid', fgColor="CCEEFF")
                
                if fontMax is not None and c in (4,6,10) and index >=1:
                    cell1.font = fontMax
                elif fontBoll is not None and c in  (4,6,8) and index >=1:
                    cell1.font = fontBoll
                else:
                    cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
                
                cell1.border = border

        self.rows = self.rows + rows
        self.AddTail(sheet)
        self.formatColumnsWidth(sheet)


class CScoreZaiMgr(object):
    def __init__(self,dbConnection,date) -> None:
        self.dbConnection = dbConnection
        self.ZhuanZaiInfo = None
        self.QuantileDf = None
        self.date = date

    def _score(self,volumn,percentail,reversed = False):
        result = 1
        for index, value in percentail.items():
            if float(volumn) <= float(value):
                result = float(index)
                break
        
        if reversed:
            result = 1.0 - float(result)
        
        result = int(result * 100)
        return result

    def GetKeZhuanZaiInfo(self):
        sql = f'''select * FROM stock.kezhuanzhai_all where `日期` = "{self.date}";'''
        results, columns = self.dbConnection.Query(sql)
        self.ZhuanZaiInfo = pd.DataFrame(results,columns=columns)
        self.ZhuanZaiInfo.set_index(["转债代码",],drop=True,inplace=True)
        self.ZhuanZaiInfo['剩余规模'] = self.ZhuanZaiInfo['剩余规模'].astype(float)

        sql = f'''SELECT stockID as `转债代码`,`delta` as `涨幅差`  FROM stock.compareindex_zai where `date`= '{self.date}';'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        df.set_index(["转债代码",],drop=True,inplace=True)
        self.ZhuanZaiInfo["涨幅差"] = df["涨幅差"]

    def _Quantile(self):
        if self.ZhuanZaiInfo is None:
            return 
        newDf = pd.DataFrame(self.ZhuanZaiInfo,columns=["成交额(万元)","剩余规模","涨幅差"])
        step = list(np.linspace(0, 1, 101))
        self.QuantileDf = newDf.quantile(step)
        print(self.QuantileDf)
        
    def CalcVolumnScore(self):
        #计算成交量分数值
        if self.ZhuanZaiInfo is None:
            return 
        
        self.ZhuanZaiInfo["成交量分数"] = self.ZhuanZaiInfo.apply(lambda row: self._score(row["成交额(万元)"],self.QuantileDf["成交额(万元)"],False), axis=1)
        
    def CalcShengyuGuiMo(self):
        #计算剩余规模分数值
        if self.ZhuanZaiInfo is None:
            return 
        
        self.ZhuanZaiInfo["剩余规模分数"] = self.ZhuanZaiInfo.apply(lambda row: self._score(row["剩余规模"],self.QuantileDf["剩余规模"],True), axis=1)
        
    def CalcStrongerThanIndexScore(self):
        # 技术比指数强的分数值，1.比指数抗跌，2 比指数涨的多
        if self.ZhuanZaiInfo is None:
            return 
        
        self.ZhuanZaiInfo["比指数分数"] = self.ZhuanZaiInfo.apply(lambda row: self._score(row["涨幅差"],self.QuantileDf["涨幅差"],False), axis=1)
        

    def WriteToDB(self):
        if self.ZhuanZaiInfo is None:
            return
        
        for index, row in self.ZhuanZaiInfo.iterrows():
            v = row.get('成交量分数',0)
            d = row.get('比指数分数',0)
            s = row.get('剩余规模分数',0)

            stockID = index
            if pd.isna(d):
                d = 0    

            total = float(v) * 0.333 + float(d)*0.333 + float(s)*0.3333
            sql = f'''REPLACE INTO `stock`.`kezhuanzai_score_everyday` (`日期`, `转债代码`, `成交量分数`,`比指数分数`,`剩余规模分数`,`总分`) VALUES ('{self.date}', '{stockID}', '{v}', '{d}',{s},{total:.1f});'''
            self.dbConnection.Execute(sql)

    def Score(self):
        self.GetKeZhuanZaiInfo()
        self._Quantile()
        self.CalcVolumnScore()
        self.CalcShengyuGuiMo()
        self.CalcStrongerThanIndexScore()
        self.WriteToDB()



class CSelectZai(object):
    def __init__(self,dbConnection,tradingDays):
        self.dbConnection = dbConnection
        self.tradingDays = tradingDays
        self.allBasicInfo = None
        self.matchedInfo = None
        self.basicInfoColumns = None

    def GetBasicInfo(self):
        sql = f'''SELECT `转债代码`,`转债名称`,`现价`,`成交额(万元)` ,`PB`,`溢价率`,`剩余年限`,`剩余规模`,`到期税前收益率` FROM stock.kezhuanzhai_all where `日期`="{self.tradingDays[-1]}";'''
        results, columns = self.dbConnection.Query(sql)
        self.basicInfoColumns = columns
        self.allBasicInfo = pd.DataFrame(results,columns = columns)

    def GetMatchedInfo(self):
        sql = f'''SELECT `转债代码`,`转债名称` FROM stock.kezhuanzhai where `日期`="{self.tradingDays[-1]}";'''
        results, columns = self.dbConnection.Query(sql)
        self.matchedInfo = pd.DataFrame(results,columns = columns)

    def _isMatched(self,stockID):
        matched = list(self.matchedInfo['转债代码'])
        if stockID in matched:
            return True
        
        return False

    def formatVolumn(self,volumn):
        ret = f'''{volumn:.2f}'''
        return ret
    
    def Select(self,param:map):
        self.GetBasicInfo()
        self.GetMatchedInfo()
        datas = []
        columns = None
        for key in param:
            v = param[key]
            startDay = v.get("startDay",None)
            endDay = v.get("endDay",None)
            if startDay is not None and endDay is not None:
                sql = f'''SELECT * FROM stock.kezhuanzai_score_everyday where `日期`>="{startDay}" and `日期`<="{endDay}";'''
            elif startDay is not None:
                sql = f'''SELECT * FROM stock.kezhuanzai_score_everyday where `日期`="{startDay}";'''
            elif endDay is not None:
                sql = f'''SELECT * FROM stock.kezhuanzai_score_everyday where `日期`="{endDay}";'''
            
            results, columns = self.dbConnection.Query(sql)
            datas.extend(results)
        
        datas = list(set(datas))
        df = pd.DataFrame(datas,columns = columns)
        res = []
        groups = df.groupby("转债代码")
        for stockID, group in groups:
            avg = group["总分"].mean()
            s = group["总分"].std()
            res.append({"转债代码":stockID,"平均分":avg,"方差":s})
        
        resDf = pd.DataFrame(res, columns=["转债代码", "平均分",'方差'])
    
        resDf["符合条件"] = resDf.apply(lambda row: self._isMatched(row['转债代码']), axis=1)
        resDf["平均分"] = resDf.apply(lambda row: self.formatVolumn(row['平均分']), axis=1)
        resDf["方差"] = resDf.apply(lambda row: self.formatVolumn(row['方差']), axis=1)
        resDf = pd.merge(resDf,self.allBasicInfo, how='inner',left_on=("转债代码",),right_on=("转债代码",))
        
        resDf.sort_values(['平均分',"溢价率"],axis=0,ascending=False,inplace=True)
        columns = self.basicInfoColumns + ["平均分","符合条件"]
        result = pd.DataFrame(resDf,columns = columns)
        

        fodler = GetZhuanZaiFolder(self.tradingDays[-1])
        fileName1 = f'''可转债评分EX_符合条件_带分数'''
        fileName2 = f'''可转债评分EX_符合条件_THS'''

        xlsxFileName = os.path.join(fodler,f"可转债评分EX_全_{self.tradingDays[-1]}.xlsx")
        xlsxFileName_Matched = os.path.join(fodler,f"可转债评分EX_符合条件的_{self.tradingDays[-1]}.xlsx")
        result.to_excel(xlsxFileName,index=False)

        matchedDf = result[result["符合条件"]]
        matchedDf.reset_index(inplace=True)
        matchedDf.to_excel(xlsxFileName_Matched,index=False)
        DataFrameToJPG(matchedDf,["转债代码","转债名称","平均分"],fodler,fileName1)
        DataFrameToJPG(matchedDf,["转债代码","转债名称"],fodler,fileName2)


    def SelectZhaiByScore(self,date):
        '''
        根据每日分数,跳跃最大的筛选
        举例: 5天前是60分，今天是 85分，差距是25分，
        '''
        sql = f'''SELECT A.`日期`,A.`转债代码`,B.`转债名称`,A.`成交量分数`,A.`比指数分数`,A.`剩余规模分数`,A.`总分`,B.`现价` FROM stock.kezhuanzai_score_everyday AS A, (SELECT * FROM `kezhuanzhai_all` where `日期` = "{date}") AS B where A.`转债代码` = B.`转债代码`;'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        df['成交量分数'] = df['成交量分数'].astype(float)
        df['比指数分数'] = df['比指数分数'].astype(float)
        df['剩余规模分数'] = df['剩余规模分数'].astype(float)
        df['总分'] = df['总分'].astype(float)
        groups = df.groupby("转债代码")
        result = []
        for stockID, group in groups:
            if group.shape[0] <3: 
                continue
            group.reset_index(drop=True,inplace=True)
            group["MA20"] = group["总分"].rolling(window=20).mean()
            # score5 = group.iloc[-5]['总分']
            # score4 = group.iloc[-4]['总分']
            score3 = group.iloc[-3]['总分']
            score2 = group.iloc[-2]['总分']
            score1 = group.iloc[-1]['总分']
            ma20 = group.iloc[-1]['MA20']

            scores = group['总分']
            max = scores.max()
            min = scores.min()
            avg = scores.mean()
            std = scores.std()

            quantile5 = scores.quantile(0.05)
            quantile25 = scores.quantile(0.25)
            quantile50 = scores.quantile(0.5)
            quantile75 = scores.quantile(0.75)
            quantile95 = scores.quantile(0.95)

            if score1>=80 and score1 - score2 >= 12 and score1 - score3 >=12:
                #print(group)
                res = {}
                res["日期"] = group.iloc[-1]['日期']
                res["转债代码"] = group.iloc[-1]['转债代码']
                res["转债名称"] = group.iloc[-1]['转债名称']
                res["现价"] = group.iloc[-1]['现价']
                res["今日分数"] = score1
                res["MA20"] = ma20
                res["BOLL上轨"] = ma20 + 2*std
                res["总分98分位"] = quantile95
                res["最高分"] = max

                res["昨日分数"] = score2
                res["前日分数"] = score3
                res["平均分"] = avg
                
                res["最低分"] = min
                res["分数标准差"] = std
                res["总分2分位"] = quantile5
                res["总分25分位"] = quantile25
                res["总分50分位"] = quantile50
                res["总分75分位"] = quantile75
                res["BOLL下轨"] = ma20 - 2*std

                result.append(res)

        df = pd.DataFrame(result)
        df["今日分数"] = df.apply(lambda row: '{:.02f}'.format(row["今日分数"]), axis=1)
        df["MA20"] = df.apply(lambda row: '{:.02f}'.format(row["MA20"]), axis=1)
        df["BOLL上轨"] = df.apply(lambda row: '{:.02f}'.format(row["BOLL上轨"]), axis=1)
        df["总分98分位"] = df.apply(lambda row: '{:.02f}'.format(row["总分98分位"]), axis=1)
        df["平均分"] = df.apply(lambda row: '{:.02f}'.format(row["平均分"]), axis=1)
        df["最高分"] = df.apply(lambda row: '{:.02f}'.format(row["最高分"]), axis=1)

        df["昨日分数"] = df.apply(lambda row: '{:.02f}'.format(row["昨日分数"]), axis=1)
        df["前日分数"] = df.apply(lambda row: '{:.02f}'.format(row["前日分数"]), axis=1)
        df["最低分"] = df.apply(lambda row: '{:.02f}'.format(row["最低分"]), axis=1)
        df["分数标准差"] = df.apply(lambda row: '{:.02f}'.format(row["分数标准差"]), axis=1)
        df["总分2分位"] = df.apply(lambda row: '{:.02f}'.format(row["总分2分位"]), axis=1)
        df["总分25分位"] = df.apply(lambda row: '{:.02f}'.format(row["总分25分位"]), axis=1)
        df["总分50分位"] = df.apply(lambda row: '{:.02f}'.format(row["总分50分位"]), axis=1)
        df["总分75分位"] = df.apply(lambda row: '{:.02f}'.format(row["总分75分位"]), axis=1)
        df["BOLL下轨"] = df.apply(lambda row: '{:.02f}'.format(row["BOLL下轨"]), axis=1)

        root = GetFuPanRoot(date)
        xlsxFileName = os.path.join(root,f"可转债评分_{date}.xlsx")
        mode='w'
        if_sheet_exists = None
        if os.path.exists(xlsxFileName) == True:
            mode='a'
            if_sheet_exists = 'overlay'

        with pd.ExcelWriter(xlsxFileName,engine='openpyxl',mode=mode,if_sheet_exists=if_sheet_exists) as excelWriter:
            write = CWriteZaiScoreToXlsx()
            write.sheetName = "可转债分数快速增加速查表"
            write.title =f'''可转债分数快速增加速查表({date})'''
            write.WriteScoreToXLSX(df,excelWriter)

        DataFrameToJPG(df,["转债代码","转债名称"],root,"转债分数速增")

            