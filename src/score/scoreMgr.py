import os
import pandas as pd

from workspace import workSpaceRoot,WorkSpaceFont
from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
from openpyxl.utils import column_index_from_string
from message.feishu.webhook_zhuanzai import SendkeZhuanZaiScore

def ConvertDataFrameToJPG(df,fullPath):
    from pandas.plotting import table
    import matplotlib.pyplot as plt
    plt.rcParams["font.sans-serif"] = [WorkSpaceFont]#显示中文字体
    high = int(0.174 * df.shape[0]+0.5)+1
    fig = plt.figure(figsize=(3, high), dpi=200)#dpi表示清晰度
    ax = fig.add_subplot(111, frame_on=False) 
    ax.xaxis.set_visible(False)  # hide the x axis
    ax.yaxis.set_visible(False)  # hide the y axis
    table(ax, df, loc='center')  # 将df换成需要保存的dataframe即可
    plt.savefig(fullPath)
    plt.close()

def DataFrameToJPG(df,columns,rootPath, fileName):
    size = df.shape[0]
    step = 80
    if size > step:
        for index in range(0,size,step):
            tmp = df.iloc[index:,]
            if index + step <= size:
                tmp = df.iloc[index:index+step,]
            fullPath = f"{rootPath}{fileName}_{int(index/step+1)}.jpg"
            print(fullPath)
            jpgDataFrame = pd.DataFrame(tmp,columns=columns)
            ConvertDataFrameToJPG(jpgDataFrame,fullPath)
    else:
        fullPath = f"{rootPath}{fileName}.jpg"
        print(fullPath)
        jpgDataFrame = pd.DataFrame(df,columns=columns)
        ConvertDataFrameToJPG(jpgDataFrame,fullPath)

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment_center = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

alignment_left = Alignment(
         horizontal='left',
         vertical='center',
         text_rotation=0,
         indent=0,
         wrapText=True
     )

#文字竖排
alignment_vertical = Alignment(
    horizontal='center', 
    vertical='center', 
    textRotation=255
    )


class CWriteScoreToXlsx(object):
    def __init__(self):
        self.rows = 0
        self.index = 1
        self.contextFontSize = 16
        self.sheetName = "可转债打分"
        self.title = "可转债打分"
        self.tailInfo = None
        self.firtLineInfo = None

    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 16
        sheet.column_dimensions['C'].width = 16
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 16
        sheet.column_dimensions['I'].width = 16
        sheet.column_dimensions['J'].width = 16

    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,column1,column2,rowHight = 32):
        mergeCell = f'{column1}{rowIndex}:{column2}{rowIndex}'
        sheet.merge_cells(mergeCell)
        if rowHight > 0:
            self.formatRowHeight(sheet,rowIndex,rowHight)
        c1 = column_index_from_string(column1)
        c2 = column_index_from_string(column2)
        for column in range(c1,c2+1):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def mergeColumn(self,sheet,column,start,end):
        mergeCell = f'{column}{start}:{column}{end}'
        sheet.merge_cells(mergeCell)
        c = column_index_from_string(column)
        for index in range(start,end+1):
            cell = sheet.cell(row = index, column = c)
            cell.border = border

    def addTitle(self,sheet):
        self.mergeRow(sheet,1,"A","J",140)
        cell = sheet.cell(1,1)
        cell.alignment = alignment_center
        cell.fill = PatternFill('solid', fgColor="009DDC")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=48, italic=False, color='FFFFFF', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddFirstLine(self,sheet):
        self.mergeRow(sheet,self.rows+1,"A","J",100)
        cell = sheet.cell(self.rows+1,1)
        cell.alignment = alignment_left
        cell.fill = PatternFill('solid', fgColor="FFFFFF")
        cell.value = f'''说明:\n1.红色字体部分说明数据没有达到平均值标准.\n2.平均值标准如下:{self.tailInfo}\n3.{self.firtLineInfo}'''
        cell.font = Font(name='宋体', size= self.contextFontSize, italic=False, color='FF0000', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def AddTail(self,sheet):
        self.mergeRow(sheet,self.rows+1,"A","J",32)
        cell = sheet.cell(self.rows+1,1)
        cell.alignment = alignment_left
        cell.fill = PatternFill('solid', fgColor="FFFFFF")
        cell.value =  f'''风险提示: 本内容仅信息分享,不构成投资建议,若以此作为买卖依据,后果自负。市场有风险,投资需谨慎！'''
        cell.font = Font(name='宋体', size= 10, italic=False, color='FF0000', bold=True)
        cell.border = border
        self.rows = self.rows + 1

    def WriteScoreToXLSX(self, scoreDF:pd.DataFrame,fullPath):
        avg1 = scoreDF["成交量分数"].mean()
        avg2 = scoreDF["抗跌分数"].mean()
        avg3 = scoreDF["领涨分数"].mean()
        avg4 = scoreDF["剩余规模分数"].mean()
        self.tailInfo = f'''成交量平均分: {avg1:.2f}  剩余规模平均分: {avg4:.2f}    抗跌平均分: {avg2:.2f}    领涨平均分: {avg3:.2f}'''
        rows = scoreDF.shape[0] + 1
        columns = scoreDF.shape[1]+1
        with pd.ExcelWriter(fullPath,engine='openpyxl',mode='w+') as excelWriter:
            df = pd.DataFrame()
            df.to_excel(excelWriter, sheet_name= self.sheetName,index=False)
            sheet = excelWriter.sheets[self.sheetName]
            self.addTitle(sheet)
            self.AddFirstLine(sheet)
            scoreDF.index = scoreDF.index +1
            scoreDF.to_excel(excelWriter, sheet_name= self.sheetName,index=True,index_label = "序号",startrow=self.rows,header=True)
            for index in range(0,rows):
                for c in range(1,columns+1):
                    cell1 = sheet.cell(self.rows+index+1,c)
                    cell1.alignment = alignment_center
                    cell1.fill = PatternFill('solid', fgColor="FFFFFF")
                    if index % 2 != 0:
                        cell1.fill = PatternFill('solid', fgColor="CCEEFF")
                    cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='000000', bold=True)
                    cell1.border = border
                    redFont = False
                    if c == 6 and index >=1:
                        data = scoreDF.iloc[index -1]["成交量分数"]
                        if data < avg1:
                            redFont = True
                    if c == 7 and index >=1:
                        data = scoreDF.iloc[index -1]["抗跌分数"]
                        if data < avg2:
                            redFont = True
                    if c == 8 and index >=1:
                        data = scoreDF.iloc[index -1]["领涨分数"]
                        if data < avg3:
                            redFont = True
                    if c == 9 and index >=1:
                        data = scoreDF.iloc[index -1]["剩余规模分数"]
                        if data < avg4:
                            redFont = True

                    if redFont:
                        cell1.font = Font(name='宋体', size=self.contextFontSize, italic=False, color='FF0000', bold=True)
            self.rows = self.rows + rows
            self.AddTail(sheet)
            self.formatColumnsWidth(sheet)


class CScoreMgr(object):
    def __init__(self,dbConnection) -> None:
        self.dbConnection = dbConnection
        self.ZhuanZaiInfo = None
        self.QuantileDf = None
        self.date = None
        self.diDianDate = None
        self.QuJianStr = ""

    def _score(self,volumn,percentail,reversed = False):
        result = 1
        for index, value in percentail.items():
            if float(volumn) <= float(value):
                result = float(index)
                break
        
        if reversed:
            result = 1.1 - float(result)
        
        result = int(result * 10)
        return result

    def GetKeZhuanZaiInfo(self,date):
        sql = f'''select * FROM stock.kezhuanzhai_all where `日期` = "{date}";'''
        results, columns = self.dbConnection.Query(sql)
        self.ZhuanZaiInfo = pd.DataFrame(results,columns=columns)
        self.ZhuanZaiInfo.set_index(["转债代码",],drop=True,inplace=True)
        self.ZhuanZaiInfo['剩余规模'] = self.ZhuanZaiInfo['剩余规模'].astype(float)

    def _Quantile(self):
        if self.ZhuanZaiInfo is None:
            return 
        newDf = pd.DataFrame(self.ZhuanZaiInfo,columns=["成交额(万元)","剩余规模"])
        self.QuantileDf = newDf.quantile([0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9])
        print(self.QuantileDf)
        
    def CalcVolumnScore(self):
        #计算成交量分数值
        if self.ZhuanZaiInfo is None:
            return 
        
        self.ZhuanZaiInfo["成交量分数"] = self.ZhuanZaiInfo.apply(lambda row: self._score(row["成交额(万元)"],self.QuantileDf["成交额(万元)"],False), axis=1)
        
    def CalcShengyuGuiMo(self):
        #计算市值分数值
        if self.ZhuanZaiInfo is None:
            return 
        
        self.ZhuanZaiInfo["剩余规模分数"] = self.ZhuanZaiInfo.apply(lambda row: self._score(row["剩余规模"],self.QuantileDf["剩余规模"],True), axis=1)
        
    def CalcStrongerThanIndexScore(self,startDay,endDay,columnName):
        # 技术比指数强的分数值，1.比指数抗跌，2 比指数涨的多
        sql = f'''SELECT stockID as `转债代码`,avg(flag) as {columnName} FROM stock.compareindex_zai where `date`>= '{startDay}' and `date` <= '{endDay}' group by stockID order by {columnName} DESC;'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        df.set_index(["转债代码",],drop=True,inplace=True)
       
        self.ZhuanZaiInfo[columnName] = df[columnName]
        self.ZhuanZaiInfo[f"{columnName}周期"] = f'''"{startDay}"-"{endDay}"'''

    def WriteToDB(self,date):
        if self.ZhuanZaiInfo is None:
            return
        
        for index, row in self.ZhuanZaiInfo.iterrows():
            v = row.get('成交量分数',0)
            dd = row.get('抗跌分数周期',"")
            #m = row['市值分数']
            d = row.get('抗跌分数',0)
            z = row.get('领涨分数',0)
            zz = row.get('领涨分数周期',"")
            s = row.get('剩余规模分数',0)

            stockID = index
            if pd.isna(z):
                z = 0
            if pd.isna(d):
                d = 0    

            total = float(v) * 0.25 + float(d)*0.25 + float(s)*0.25 + float(z)*0.25     
            sql = f'''REPLACE INTO `stock`.`kezhuanzai_score` (`日期`, `转债代码`, `成交量分数`, `抗跌分数周期`,`抗跌分数`, `领涨分数周期`,`领涨分数`,`剩余规模分数`,`总分`) VALUES ('{date}', '{stockID}', '{v}', '{dd}','{d}', '{zz}','{z}',{s},{total});'''
            #print(sql)
            self.dbConnection.Execute(sql)

    def Score(self,indexParams:map):
        self._preprocessing(indexParams)
        self.GetKeZhuanZaiInfo(self.date)
        self._Quantile()
        self.CalcVolumnScore()
        self.CalcShengyuGuiMo()
        for key in indexParams:
            startDay = indexParams[key]["startDay"]
            endDay = indexParams[key]["endDay"]
            self.CalcStrongerThanIndexScore(startDay,endDay,key)

        self.WriteToDB(self.date)
    

    def ConverBigVolumnOfZhuanZhaiToJEPG(self, date,res):
        result = []
        for r in res:
            zhuanZaiDaiMa = r["转债代码"]
            zhuanZai_name = r["转债名称"]
            stockID = r["股票代码"]
            stockName = r["股票简称"]
            dict1 = {"代码": zhuanZaiDaiMa,"名称":zhuanZai_name}
            dict2 = {"代码": stockID,"名称":stockName}
            result.append(dict1)
            result.append(dict2)

        jpgDataFrame = pd.DataFrame(result, columns=("代码","名称"))
        folderRoot= f'''{workSpaceRoot}/复盘/可转债评分/{date}/'''
        if os.path.exists(folderRoot) == False:
            os.makedirs(folderRoot)

        DataFrameToJPG(jpgDataFrame,("代码","名称"),folderRoot,f"可转债放量")


    def _preprocessing(self,indexParam):
        self.QuJianStr = ""
        self.date = None
        self.diDianDate = None

        dates = []
        if "领涨分数" in indexParam:
            dates.append(indexParam ["领涨分数"]["endDay"])
            self.diDianDate = indexParam ["领涨分数"]["startDay"]
            
        if "抗跌分数" in indexParam:
            d = indexParam ["抗跌分数"]["endDay"]  
            dates.append(d)
            self.diDianDate = d   #最低点的时候价格

        maxDate = max(dates)
        sql = f"SELECT `日期` FROM stock.treadingDay where `开市` =1 and `交易所`='SSE' and `日期`>='{maxDate}'"
        res,_ = self.dbConnection.Query(sql)
        results = [r[0] for r in res]
        self.date = results[0]

        for key in indexParam:
            s = indexParam[key]["startDay"]
            e = indexParam[key]["endDay"]
            if key == "领涨分数":
                self.QuJianStr = self.QuJianStr + f'''上涨区间: {s}  -  {e}      '''
            elif key == "抗跌分数":
                self.QuJianStr = self.QuJianStr + f'''下跌区间: {s}  -  {e}      '''

    def _BothStockAndZhuanZaiToJPEG(self,df,folder,name):
        result = []
        for _, row in df.iterrows():
            zhuanZaiDaiMa = row["转债代码"]
            zhuanZai_name = row["转债名称"]
            stockID = row["正股代码"]
            stockName = row["正股名称"]
            dict1 = {"代码": zhuanZaiDaiMa,"名称":zhuanZai_name}
            dict2 = {"代码": stockID,"名称":stockName}
            result.append(dict1)
            result.append(dict2)
        jpgDataFrame = pd.DataFrame(result, columns=("代码","名称"))
        jpgDataFrame.index = jpgDataFrame.index + 1
        DataFrameToJPG(jpgDataFrame,("代码","名称"),folder,name)

    def Select(self,indexParams:map):
        self._preprocessing(indexParams)
        sql = f'''select A.`转债代码`,B.`转债名称`,B.`正股代码`,B.`正股名称`,A.`成交量分数`,A.`抗跌分数`,A.`领涨分数`,A.`剩余规模分数`,A.`总分`from kezhuanzai_score As A,kezhuanzhai AS B where A.`日期` = "{self.date}" and B.`日期` = "{self.diDianDate}"  and A.`转债代码` = B.`转债代码` order by A.`总分` DESC;'''
        results, columns = self.dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        avg1 = df["成交量分数"].mean()
        avg2 = df["抗跌分数"].mean()
        avg3 = df["领涨分数"].mean()
        avg4 = df["剩余规模分数"].mean()
        # sql2 = f'''select A.`转债代码`,B.`转债名称`,B.`正股代码`,B.`正股名称`,A.`成交量分数`,A.`抗跌分数`,A.`领涨分数`,A.`剩余规模分数`,A.`总分`from kezhuanzai_score As A,kezhuanzhai AS B where A.`日期` = "{self.date}"  and (A.`成交量分数`<= {avg1:.2f} or A.`抗跌分数`<= {avg2:.2f} or A.`领涨分数`<= {avg3:.2f} or A.`剩余规模分数`<= {avg4:.2f}) and B.`日期` = "{self.diDianDate}"  and A.`转债代码` = B.`转债代码` order by A.`总分` DESC;'''
        # print(sql2)
        newDF = pd.DataFrame(df)
        newDF = newDF[newDF["成交量分数"] > avg1]
        newDF = newDF[newDF["抗跌分数"] > avg2]
        newDF = newDF[newDF["领涨分数"] > avg3]
        newDF = newDF[newDF["剩余规模分数"] > avg4]
        newDF.reset_index(drop=True,inplace=True)
        newDF.index = newDF.index + 1

        fodler = f'{workSpaceRoot}/复盘/可转债评分/{self.date}/'
        if os.path.exists(fodler) == False:
            os.makedirs(fodler)
        fileName1 = f'''可转债评分'''
        fileName2 = f'''可转债评分_完全符合条件'''
        DataFrameToJPG(df,["转债代码","转债名称","总分"],fodler,fileName1)
        DataFrameToJPG(newDF,["转债代码","转债名称"],fodler,fileName2)

        xlsxFileName = os.path.join(fodler,"可转债评分.xlsx")
        xlsx = CWriteScoreToXlsx()
        xlsx.title = f'''可转债评分表 ({self.date})'''
        xlsx.firtLineInfo = self.QuJianStr
        xlsx.WriteScoreToXLSX(df,xlsxFileName)
        self._BothStockAndZhuanZaiToJPEG(df,fodler,"可转债评分_带正股")
        self._BothStockAndZhuanZaiToJPEG(newDF,fodler,"可转债评分_完全符合条件_带正股")
        webhook = "https://open.feishu.cn/open-apis/bot/v2/hook/e156ab0d-9d9d-4bc4-a4b5-faf9ad6344c2"
        secret = "chzCzY4VkzctfN2qvtxARg"
        # webhook = "https://open.feishu.cn/open-apis/bot/v2/hook/4901573e-b858-434a-a787-5faa28982b1a" # 测试API
        # secret = "brYyzPbSks4OKnMgdwKvIh"
        SendkeZhuanZaiScore(self.dbConnection,self.date,self.diDianDate,webhook,secret,50)


            