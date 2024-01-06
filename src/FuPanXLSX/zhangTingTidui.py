from openpyxl.styles import Font,Border,Side,Alignment,Font,PatternFill
import pandas as pd

# 边框
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

# 对齐
alignment = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0
     )

class CWriteZhangTingTiDuiToXLSX(object):
    def __init__(self,date):
        self.rows = 0
        self.date = date
        self.sheetName = f'''涨停梯队'''
        self.title = f'''涨停梯队复盘表 ({self.date})'''
    
    def formatColumnsWidth(self,sheet):
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 85
    
    def formatRowHeight(self,sheet,rowIndex,height):
        sheet.row_dimensions[rowIndex].height=height

    def mergeRow(self,sheet,rowIndex,rowHight = 32):
        mergeCell = f'A{rowIndex}:H{rowIndex}'
        sheet.merge_cells(mergeCell)
        self.formatRowHeight(sheet,rowIndex,rowHight)
        for column in range(1,9):
            cell = sheet.cell(row = rowIndex, column = column)
            cell.border = border

    def addTitle(self,sheet):
        cell = sheet.cell(1,1)
        cell.alignment = alignment
        cell.fill = PatternFill('solid', fgColor="B5C6EA")
        cell.value =  self.title
        cell.font = Font(name='宋体', size=28, italic=False, color='000000', bold=True)
        cell.border = border
        self.mergeRow(sheet,1,40)
        self.rows = self.rows + 1

    def formatGaiNianStyle(self,cell):
        cell.font = Font(name='宋体', size=22, italic=False, color='FF0000', bold=True)
        cell.border = border
        cell.alignment = alignment
        cell.fill = PatternFill('solid', fgColor="FFF3CA")
    

    def formatContextStyle(self,cell):
        cell.font = Font(name='宋体', size=14, italic=False, color='000000', bold=True)
        cell.border = border
        cell.alignment = alignment

    def formatAllContext(self,sheet,lines):
        startIndex = self.rows + 1
        endIndex = self.rows + lines
        for line in range(startIndex,endIndex+1):
            for column in range(1,9):
                cell = sheet.cell(row = line, column = column)
                self.formatContextStyle(cell)

    def WriteZhangTingTiDuiToXLSX(self,datas,excelWriter:pd.ExcelWriter):
        tmp = pd.DataFrame()
        tmp.to_excel(excelWriter, sheet_name= self.sheetName)
        sheet = excelWriter.sheets[self.sheetName]
        self.addTitle(sheet)
        for index,data in enumerate(datas):
            gaiNian,df = data
            if index == 0:
                self.mergeRow(sheet,2,32)
                cell = sheet.cell(2,1)
                cell.value = f'''"{gaiNian}" 概念'''
                self.formatGaiNianStyle(cell)

                self.rows = 2
                df.to_excel(excelWriter, sheet_name=self.sheetName, index=False,startrow=self.rows,header=True)
                lines = df.shape[0] + 1  # 因为有表头
                self.formatAllContext(sheet,lines)
                self.rows = self.rows + df.shape[0] + 4 # 因为有表头
            else:
                self.mergeRow(sheet,self.rows,32)
                cell = sheet.cell(self.rows,1)
                cell.value = f'''"{gaiNian}"概念'''

                self.formatGaiNianStyle(cell)
                df.to_excel(excelWriter, sheet_name=self.sheetName, index=False,startrow=self.rows,header=False)
                lines = df.shape[0]  # 没有表头
                self.formatAllContext(sheet,lines)
                self.rows = self.rows + df.shape[0] + 3
            #标题
            self.rows = self.rows + 2
            self.formatColumnsWidth(sheet)

    def AnalysisZhangTingReason(self,dbConnection,excelWriter:pd.ExcelWriter):
        sql = f"select A.*,B.`所属概念`from `stockZhangting`AS A, `stockBasicInfo` As B where A.`股票代码`= B.`股票代码` and `日期` = '{self.date}' order by `连续涨停天数` DESC,`首次涨停时间` ASC;"
        results, columns = dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        reasons = []
        for _, row in df.iterrows():
            reason = row["涨停原因类别"]
            reasons.extend(reason.split("+"))
    
        reasons = list(set(reasons))
        reasonResults = {}
        for reason in reasons:
            sql = f'''select A.`日期`,A.`股票代码`,A.`股票简称`,A.`连续涨停天数` as `涨停天数`,A.`涨停关键词` as `连板数`,A.`首次涨停时间`,A.`最终涨停时间`,A.`涨停原因类别` as `涨停原因` from `stockZhangting`AS A, `stockBasicInfo` As B where A.`股票代码`= B.`股票代码` and `日期` = '{self.date}' and A.`涨停原因类别` like '%{reason}%' order by `连续涨停天数`DESC ,`最终涨停时间` ASC;'''
            result ,columns = dbConnection.Query(sql)
            df = pd.DataFrame(result, columns = columns)
            count = df.shape[0]
            if count >=3:
                reasonResults[reason] = (count,df)
        
        ret = sorted(reasonResults.items(), key=lambda d: d[1][0],reverse=True)
        dataFrames = [(item[0],item[1][1]) for item in ret if item[1][0]>=3]
        self.WriteZhangTingTiDuiToXLSX(dataFrames,excelWriter)

    
    def WriteZhangTingXLSX(self,dbConnection,excelWriter):
        sheetName = f"涨停复盘"
        sql = f'''SELECT `日期`,`股票代码`,`股票简称`,`连续涨停天数` as `涨停天数`,`涨停关键词` as `连板数`,`首次涨停时间`,`最终涨停时间`,`涨停原因类别` as `涨停原因`  FROM stock.stockzhangting where `日期` = "{self.date}" order by `连续涨停天数` DESC ,`首次涨停时间` ASC , `涨停关键词` DESC ;'''
        results, columns = dbConnection.Query(sql)
        df = pd.DataFrame(results,columns=columns)
        df.to_excel(excelWriter, sheet_name= sheetName,index=False,startrow=1)
        sheet = excelWriter.sheets[sheetName]
        cell = sheet.cell(1,1)
        cell.alignment = alignment
        cell.fill = PatternFill('solid', fgColor="B5C6EA")
        cell.value =  f"涨停复盘 {self.date}"
        cell.font = Font(name='宋体', size=28, italic=False, color='000000', bold=True)
        cell.border = border
        self.mergeRow(sheet,1,40)
        startRow = 1
        endRow = startRow + df.shape[0] +1
        for index in range(1,endRow):
            for column in range(1,9):
                cell = sheet.cell(row = startRow + index, column = column)
                cell.border = border
                cell.alignment = alignment
                cell.fill = PatternFill('solid', fgColor="000000")
                if index % 2 != 0:
                    cell.fill = PatternFill('solid', fgColor="CCEEFF")
                cell.font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)
                cell.border = border


        self.formatColumnsWidth(sheet)